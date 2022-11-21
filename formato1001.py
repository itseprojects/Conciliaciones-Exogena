# -*- coding: utf-8 -*-
"""
Created on Fri Oct 28 17:19:31 2022

@author: HaroldFerneyGomez
"""
from cmath import nan
import logging
from operator import index

from openpyxl import load_workbook

# from openpyexcel import load_workbook
# import azure.functions as func

from datetime import datetime, timedelta
import pandas as pd
# from azure.storage.blob import BlobSasPermissions, generate_blob_sas, BlobServiceClient
import urllib.parse
from math import ceil
from io import BytesIO

account_name = 'probepython'
account_key = 'Zas0npJX9ryEm4hmW/gatWr8aI91oOCvt+qbQKqWrZJCmhv5qh6S/w6ittYYaDBRjRnoxa0h+A8H+ASttcvrrQ=='
container_name = 'test'

blob_name_DB = 'BASE DE DATOS EXELTIS.xls'
HeaderHojaDB = 0
nombreHojaDB="Sheet1"

def main():#(req: func.HttpRequest) -> func.HttpResponse:
    # blob_name = req.get_json().get('Balance')
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1"
    # ColumnaValorIngreso = "2022/08" if blob_name=='BALANCE ENERO A AGOSTO 2022.xlsx' else "2021/12"
    # blob_name_to_save = 'HG_'+blob_name
    # logging.info('Python HTTP trigger function processed a request.')
    # message = req.get_json().get('Balance')
    try:
        # blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
        # try:
        #     blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_to_save)
        #     blob_client.delete_blob()
        # except:
        #     pass
        # blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name)
        # downloader = blob_client.download_blob()
        #Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=HeaderHojaBalance)

        Datos = pd.read_excel(FilePath, sheet_name=nombreHojaBalance, nrows=0, header=10)    
        ColumnaValorIngreso = Datos.columns.get_loc("Saldo final a")
        Datos = pd.read_excel(FilePath, sheet_name=nombreHojaBalance, header=HeaderHojaBalance)
        Datos = Datos[~Datos['Cuentas'].isnull()]
        DatosSeparados = separarCuentas(Datos)
        ColumnaValorIngreso += 3
        TercerosPorConcepto = pd.DataFrame()
        # print(TercerosPorConcepto.columns)
        TercerosPorConcepto['Concepto'] = None
        TercerosPorConcepto['Numero de identificacion']= None
        TercerosPorConcepto['Razón social'] = None
        TercerosPorConcepto['Ingresos Brutos Recibidos'] = None
        
        # print(TercerosPorConcepto.columns)
        for clave in conceptos:
            print("**************************************************  "+clave)
            estadoiteracion=True
            soloConcepto = pd.DataFrame()
            soloConcepto['Numero de identificacion']= None
            soloConcepto['Razón social'] = None
            soloConcepto['Ingresos Brutos Recibidos'] = None
            # print(soloConcepto.columns)
            for conc in conceptos[clave]:
                print(">> "+str(conc))
                if str(conc).startswith("15"):
                    data1 = UnificarClientesPorCuenta(DatosSeparados,158800,159000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                    if data1.empty: 
                        data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ 1000),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                    elif estadoiteracion:
                        estadoiteracion=False
                        data=data1
                    else: data=pd.DataFrame()
                elif str(conc).startswith("16"):
                    if str(conc).startswith("1699"):
                        data = UnificarClientesPorCuenta(DatosSeparados,168800,169000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                    else:
                        data1 = UnificarClientesPorCuenta(DatosSeparados,168800,169000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                        if data1.empty: 
                            data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ 1000),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                        elif estadoiteracion:
                            estadoiteracion=False
                            data=data1
                        else: data=pd.DataFrame()
                else: data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ (100 if str(conc).endswith("00") else 1)),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                # print(data.columns)
                soloConcepto = pd.concat([soloConcepto, data], ignore_index=True ).groupby(['Numero de identificacion','Razón social']).sum().reset_index()
                    
                # print((soloConcepto))
            if ~soloConcepto.empty: 
                soloConcepto.insert(0,'Concepto', clave)
                TercerosPorConcepto = pd.concat([TercerosPorConcepto, soloConcepto],ignore_index=True)
                print(soloConcepto)
        print(TercerosPorConcepto[['Razón social','Ingresos Brutos Recibidos']])
        return TercerosPorConcepto       

        
    
        Totales = ObtenerIngresos(DatosSeparados,ColumnaValorIngreso)
        IngresosClientes4135 = UnificarClientesPorCuenta(DatosSeparados,413500,417000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
        IngresosClientes4175 = UnificarClientesPorCuenta(DatosSeparados,417500,417600,"Devoluciones",ColumnaValorIngreso)
        IngresosClientes4001 = pd.merge(IngresosClientes4135,IngresosClientes4175,on=['Numero de identificacion','Razón social']).fillna(0)
        IngresosClientes4001.insert(0,'Concepto', 4001)
        IngresosClientes4002 = UnificarClientesPorCuenta(DatosSeparados,420000,430000,"Ingresos Brutos Recibidos",ColumnaValorIngreso).fillna(0)
        IngresosClientes4002.insert(0,'Concepto', 4002)
        IngresosClientes4002.insert(2,'Devoluciones', 0)
        Ingresos = pd.concat([IngresosClientes4001,IngresosClientes4002])
        Ingresos = Ingresos.sort_values(by = ['Concepto','Ingresos Brutos Recibidos'], ascending = [True, True], na_position = 'last',ignore_index=True)
       
        ValorTotalIngresos= Ingresos["Ingresos Brutos Recibidos"].sum()+Ingresos["Devoluciones"].sum()
        print(ValorTotalIngresos)
        # blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_DB)
        # downloader = blob_client.download_blob()
        # BD = pd.read_excel(downloader.readall(), sheet_name=nombreHojaDB, header=HeaderHojaDB)
        
        # Formato1007 = BuscarId(Ingresos,BD)
        # abs_container_client = blob_service_client.get_container_client(container=container_name)
        # buffer = BytesIO()
        # excel_buf = Formato1007.to_excel(sheet_name="Formato1007",excel_writer=buffer,index=False)

        # blockBlob = abs_container_client.upload_blob(name=blob_name_to_save,data=buffer.getvalue())

    except Exception as e:
        print("error")
    #     logging.info(e)

    #     return func.HttpResponse(
    #             f"!! Ocurrió un error en la ejecución. \n\t {e} ",
    #             status_code=200
    #     )
    # return func.HttpResponse(str(ValorTotalIngresos))

def BuscarId(dfBalance,bd):
    TiposDoc={'C':13,'E':31,'N':31,'O':43,'X':31}
    ListadoIds = dfBalance["Numero de identificacion"].tolist()
    vectorCoincidencias = []
    vectorPrimerApellido = []
    vectorSegundoApellido = []
    vectorNombre = []
    vectorOtrosNombres = []
    bd['Código']=bd['Código'].apply(lambda x: x.strip())
    for Id in ListadoIds:
        try:
            tipoId = bd[(bd['Código']==str(Id))]['Tipo de identificación'].iloc[0]
            NumeroTipoID = TiposDoc[tipoId]
            if NumeroTipoID == 13:
                nombrecompleto = dfBalance[(dfBalance["Numero de identificacion"]==str(Id))]['Razón social'].iloc[0]
                if len(nombrecompleto.split())==1:
                    vectorPrimerApellido.append(nombrecompleto)
                    vectorSegundoApellido.append("")
                    vectorNombre.append("")
                    vectorOtrosNombres("")
                elif len(nombrecompleto.split())==2:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append("")
                    vectorOtrosNombres("")
                elif len(nombrecompleto.split())==3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())>3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append(nombrecompleto.split(maxsplit=3)[3])
            else:
                vectorPrimerApellido.append("")
                vectorSegundoApellido.append("")
                vectorNombre.append("")
                vectorOtrosNombres.append("")
            vectorCoincidencias.append(TiposDoc[tipoId])
        except Exception:
            vectorCoincidencias.append("NE")  
            vectorPrimerApellido.append("")
            vectorSegundoApellido.append("")
            vectorNombre.append("")
            vectorOtrosNombres.append("")
    dfBalance.insert(1,'Tipo de documento', vectorCoincidencias)
    dfBalance.insert(4,'Pais', 169)
    dfBalance.insert(3,'Otros nombres', vectorOtrosNombres)
    dfBalance.insert(3,'Primer nombre', vectorNombre)
    dfBalance.insert(3,'Segundo apellido', vectorSegundoApellido)
    dfBalance.insert(3,'Primer apellido', vectorPrimerApellido)
    return dfBalance

def separarCuentas(df):
    cuentas=df["Cuentas"]
    if "Descripción" in cuentas[0]:
        cuentas=cuentas[1:]
        df = df[1:]
    VectorCuentas = []
    VectorNits = []
    VectorNombres = []
    for cuenta in cuentas:  
        try:
            nit,name = cuenta.split(maxsplit=1)
        except Exception:
            try:
                cta = int(str(cuenta).strip().replace('.', ''))
            except Exception:
                cta = 0
            name = ""
            nit = ""
        VectorNombres.append(name)
        VectorCuentas.append(cta)
        VectorNits.append(nit)
    df.insert(1,"Razón social",VectorNombres)
    df.insert(1,"NIT",VectorNits)
    df.insert(1,"NumeroCuenta",VectorCuentas)
    return df

def ObtenerIngresos(df,ColumnaValorIngreso):
    ingresosAO = df[(df['NumeroCuenta']>413499)&(df['NumeroCuenta']<417000)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('4135+55',': ',ingresosAO)#,df[(df['NumeroCuenta']==4135)].iloc[0]['             Descripción                                '].strip()
    devoluciones= df[(df['NumeroCuenta']<417600)&(df['NumeroCuenta']>417499)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('4175',': ',devoluciones)#df[(df['NumeroCuenta']==4175)].iloc[0]['             Descripción                                '].strip(),
    ingresosO = df[(df['NumeroCuenta']<430000)&(df['NumeroCuenta']>420999)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('42',': ',ingresosO)#df[(df['NumeroCuenta']==42)].iloc[0]['             Descripción                                '].strip(),
    print('TOTAL INGRESOS :',ingresosAO+devoluciones+ingresosO)
    return ingresosAO+devoluciones+ingresosO

def UnificarClientesPorCuenta(df,limiteInferiorCta,LimiteSuperiorCta,nombreDatoColumna,ColumnaValorIngreso):
    datosPorCliente = pd.DataFrame()
    listaSoloClientesIngreso = df[(df['NIT']!="")&(df['NumeroCuenta']>=limiteInferiorCta)&(df['NumeroCuenta']<LimiteSuperiorCta)]
    listaUnica = listaSoloClientesIngreso["Razón social"].unique().tolist()
    VectorNombres = []
    VectorIdentificacion = []
    VectorTotales = []
    for cliente in listaUnica:
        total = listaSoloClientesIngreso[(listaSoloClientesIngreso['Razón social']==cliente)].iloc[:, ColumnaValorIngreso].sum()
        # print(cliente,total)
        VectorIdentificacion.append(df[df['Razón social']==cliente]['NIT'].iloc[0])
        VectorNombres.append(cliente)
        VectorTotales.append(total)
        
    datosPorCliente.insert(0,"Razón social",VectorNombres)
    datosPorCliente.insert(0,"Numero de identificacion",VectorIdentificacion)
    datosPorCliente.insert(2,nombreDatoColumna,VectorTotales)
    print(limiteInferiorCta," hasta limite ",LimiteSuperiorCta,datosPorCliente)
    return datosPorCliente

def GuardarExcel(df, nombreHoja):
    ExcelWorkbook = load_workbook(FilePath)
    writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
    writer.book = ExcelWorkbook
    df.to_excel(writer, sheet_name = nombreHoja ,index=False)
    writer.save()
    writer.close()
   
# si termina en dos ceros es de rango completo ej: 515500 va hasta 516000
conceptos = {"5055":[515500,525500,725500], 
            "5056":[519520],
            "5002":[511000,521000,721000],
            "5003":[530516],
            "5004":[513000,513500,514000,514500,515000,523500,523600,524000,524500,530505,722500,723000,723500,724000,724500],
            "5005":[512000,722000,169920],
            "5006":[530520,530521,530522,530523],
            "5063":[531521],
            "5007":[141000,143500,144000],
            "5008":[158800,150800,152400,152500,152800,154000],
            "5010":[510565,510568,510571,520572,520575,520578],
            "5011":[510574,510577],
            "5012":[510580],
            "5013":[],
            "5014":[],
            "5015":[521505,521508,530550,721500],
            "5066":[],
            "5058":[512500,522500],
            "5060":[],
            "5016":[161000,163000,164000,166000,517000,519525,519530,519595,519900,529510,529530,529540,529560,529595,530525,530526,530540,532010,539500,727000,729500],
            "5020":[],
            "5027":[],
            "5023":[],
            "5067":[],
            "5024":[],
            "5025":[],
            "5026":[],
            "5080":[],
            "5081":[],
            "5082":[],
            "5083":[],
            "5084":[],
            "5085":[],
            "5028":[],
            "5029":[],
            "5030":[],
            "5031":[],
            "5032":[169900,179900,189900,199900],
            "5033":[],
            "5034":[],
            "5035":[],
            "5019":[516500,726500],
            "5044":[],
            "5046":[],
            "5045":[],
            "5059":[],
            "5061":[],
            "5068":[],
            "5069":[],
            "5070":[],
            "5071":[],
            "5073":[],
            "5074":[],
            "5075":[],
            "5076":[],
            "5079":[],
             }
ivaColumnaO = [511535,521570]


def ExtraerCtas():
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1"
    
    try:
        Datos = pd.read_excel(FilePath, sheet_name=nombreHojaBalance, header=HeaderHojaBalance)
        Datos = Datos[~Datos["             Descripción                                "].isnull()]
        Datos1 = Datos.iloc[:,[0,1]]
        print(Datos1)
        return Datos1
    except : False

FilePath = "balance 2021 Exeltis con terceros.xlsx"
# df = ExtraerCtas()
# GuardarExcel(df, 'soloCtas')
df = main()
# print(df)
# GuardarExcel(df, 'Conceptos1001')

# conceptos = {"5055":[5155,5255,7255],
#              "5056":[],
#              "5002":[5110,5210,7210],
#              "5003":[],
#              "5004":[5235,5236,5240,5245,7225,7230,7235,7240,7245],#5135?
#              "5005":[5120,5225,7220],
#              "5006":[],
#              "5063":[],
#              "5007":[1410,1435,1440,1490],
#              "5008":[1508,1524,1525,1528,1540,1588],
#              "5010":[],
#              "5011":[],
#              "5012":[],
#              "5013":[],
#              "5014":[],
#              "5015":[5115,5215,7215],
#              "5066":[],
#              "5058":[],
#              "5060":[],
#              "5016":[1610,1630,1660,1699,5170,5195,5199,5295,5305,5315,5320,5395,7270,7295],
#              "5020":[],
#              "5027":[],
#              "5023":[],
#              "5067":[],
#              "5024":[],
#              "5025":[],
#              "5026":[],
#              "5080":[],
#              "5081":[],
#              "5082":[],
#              "5083":[],
#              "5084":[],
#              "5085":[],
#              "5028":[],
#              "5029":[],
#              "5030":[],
#              "5031":[],
#              "5032":[],
#              "5033":[],
#              "5034":[],
#              "5035":[],
#              "5019":[5165,7265],
#              "5044":[],
#              "5046":[],
#              "5045":[],
#              "5059":[],
#              "5061":[],
#              "5068":[],
#              "5069":[],
#              "5070":[],
#              "5071":[],
#              "5073":[],
#              "5074":[],
#              "5075":[],
#              "5076":[],
#              "5079":[],
#              }


