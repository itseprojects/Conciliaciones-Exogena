from cmath import nan
import logging
from operator import index

import azure.functions as func

from datetime import datetime, timedelta, date
import pandas as pd
from azure.storage.blob import BlobSasPermissions, generate_blob_sas, BlobServiceClient
import urllib.parse
from math import ceil
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
import numpy as np
import json

# Datos del blob storage
account_name = 'itseblobdev' #'probepython'
account_key = 'd9sOh0WeqvVF66NQnyWKZWFL/KDje0LizX8UyFWpWX39lLX2C8fxnqRtYD2lOFvNp6aaayQsAq7T+AStvsHyew=='

blob_name_DB = 'BASE DE DATOS EXELTIS.xls'
HeaderHojaDB = 0
nombreHojaDB="Sheet1"
blob_name_to_save = 'Formato1007-'+ str(date.today())+'.xlsx' 

def main(req: func.HttpRequest) -> func.HttpResponse:
    Cliente = req.get_json().get('Cliente') # cliente1
    TipoBalance = req.get_json().get('TipoBalance') #"SIESA"
    blob_name_DB = req.get_json().get('BaseDeDatos') #'BASE DE DATOS EXELTIS.xls'
    balanceFile = req.get_json().get('Balance')

    if TipoBalance=="SIESA":
        # Datos para la base de datos
        # blob_name_DB = 'BASE DE DATOS EXELTIS.xls'
        HeaderHojaDB = 0
        nombreHojaDB="Sheet1"
        exito = WorkSiesa(Cliente,balanceFile,blob_name_DB)
    else: 
        dicToReturn = {"error":"Tipo de balance no implementado"}
        exito = json.dumps(dicToReturn) 

    return func.HttpResponse(f"{exito}", status_code=200 )

def WorkSiesa(container_name,balanceFile,blob_name_DB):
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1"
    # ColumnaValorIngreso = "2022/08" if balanceFile=='BALANCE ENERO A AGOSTO 2022.xlsx' else "2021/12"
    try:
        blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
        try:
            blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_to_save)
            blob_client.delete_blob()
        except:
            pass
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, nrows=0, header=10,engine='openpyxl')  
        ColumnaValorIngreso = Datos.columns.get_loc("Saldo final a")
        
        # leer resto del balance
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=HeaderHojaBalance,engine='openpyxl')
        Datos = Datos[~Datos['Cuentas'].isnull()]
        DatosSeparados = separarCuentas(Datos)
        ColumnaValorIngreso += 3
        Totales = ObtenerIngresos(DatosSeparados,ColumnaValorIngreso)
        IngresosClientes4135 = UnificarClientesPorCuenta(DatosSeparados,413500,417000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
        IngresosClientes4175 = UnificarClientesPorCuenta(DatosSeparados,417500,417600,"Devoluciones",ColumnaValorIngreso)
        IngresosClientes4001 = pd.merge(IngresosClientes4135,IngresosClientes4175,on=['Numero de identificacion','Razón social'],how='outer').fillna(0)
        IngresosClientes4001.insert(0,'Concepto', 4001)
        IngresosClientes4002 = UnificarClientesPorCuenta(DatosSeparados,420000,430000,"Ingresos Brutos Recibidos",ColumnaValorIngreso).fillna(0)
        IngresosClientes4002.insert(0,'Concepto', 4002)
        IngresosClientes4002.insert(2,'Devoluciones', 0)
        Ingresos = pd.concat([IngresosClientes4001,IngresosClientes4002])
        Ingresos = Ingresos.sort_values(by = ['Concepto','Ingresos Brutos Recibidos'], ascending = [True, True], na_position = 'last',ignore_index=True)
       
        ValorTotalIngresos= Ingresos["Ingresos Brutos Recibidos"].sum()+Ingresos["Devoluciones"].sum()
        
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_DB)
        downloader = blob_client.download_blob()
        BD = pd.read_excel(downloader.readall(), sheet_name=nombreHojaDB, header=HeaderHojaDB)
        
        Formato1007 = BuscarId(Ingresos,BD)
        Formato1007["Razón social"]= np.where(Formato1007["Primer apellido"]!="","",Formato1007["Razón social"])
        Formato1007["Pais"]= np.where(Formato1007["Tipo de documento"]!=13,"",Formato1007["Pais"])
        Formato1007["Ingresos Brutos Recibidos"]= Formato1007["Ingresos Brutos Recibidos"]*-(1)
        Formato1007 = Formato1007.drop(Formato1007[(Formato1007['Ingresos Brutos Recibidos']==0) & (Formato1007['Devoluciones']==0)].index)
        PutColorsAnsSaveToBlob(Formato1007,container_name)

    except Exception as e:
        dicToReturn = {"error":f'{e}'}
        return json.dumps(dicToReturn)
    dicToReturn = {
        "error":"ninguno",
        "ruta":f'https://{account_name}.blob.core.windows.net/{container_name}/{blob_name_to_save}',
        "valorIngresos":ceil(ValorTotalIngresos*-1)}
    return json.dumps(dicToReturn)

# Función que agrega color según el tipo de contenido de las columnas y guarda archivo excel en el Blob Storage,
#   Asigna el color rojo para advertir sobre datos relacionados de la base de datos de usuarios y amarillo para datos relacionados con el balance
#   requiere de: 
#       Datos = df del formato con todas las columnas
#       account_name, account_key, blob_name_to_save
#   retorna: none
def PutColorsAnsSaveToBlob(Datos,container_name):
    Datos['Ingresos Brutos Recibidos'] = Datos['Ingresos Brutos Recibidos'].apply(np.ceil)
    Datos['Devoluciones'] = Datos['Devoluciones'].apply(np.ceil)
    
    fillOrange = PatternFill(patternType='solid', fgColor='FCBA03')
    fillRed = PatternFill(patternType='solid', fgColor='EE1111')

    nombreHojaBalance= "iva+Names" # Nombre de la hoja a guardar
    blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
    buffer = BytesIO()
    excel_buf = Datos.to_excel(sheet_name=nombreHojaBalance,excel_writer=buffer,index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(Datos, index=False, header=True):
        ws.append(r)
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    table = openpyxl.worksheet.table.Table(ref='A1:K'+str(ws.max_row), displayName='Formato1007ByHG', tableStyleInfo=mediumStyle)
    ws.add_table(table)
    
    # Agrega color a las celdas segun las siguientes normas
    for row, rowval in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in rowval:
            
            # rojo si Tipo de doc = NE, no encontrado en bd usuarios
            if (f"{cell.column_letter}"=="B" or f"{cell.column_letter}"=="I") and (f"{cell.value}"=="NE" or f"{cell.value}"=="" or f"{cell.value}"=="None" or f"{cell.value}"=="nan"):
                ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
                
            # Naranja si Ivas o Retenciones son negativas
            elif (   f"{cell.column_letter}"=="J" or f"{cell.column_letter}"=="K" ): 
                if cell.value<0 :
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillOrange
                    
    # Subir excel al blob Storage
    abs_container_client = blob_service_client.get_container_client(container=container_name)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        blockBlob = abs_container_client.upload_blob(name=blob_name_to_save,data=stream)
    return None


def BuscarId(dfBalance,bd):
    TiposDoc={'C':13,'E':31,'N':31,'O':43,'X':31}
    ListadoIds = dfBalance["Numero de identificacion"].tolist()
    vectorCoincidencias = []
    vectorPrimerApellido = []
    vectorSegundoApellido = []
    vectorNombre = []
    vectorOtrosNombres = []
    bd['Código']=bd['Código'].apply(lambda x: x.strip())
    for Id in ListadoIds:
        try:
            tipoId = bd[(bd['Código']==str(Id))]['Tipo de identificación'].iloc[0]
            NumeroTipoID = TiposDoc[tipoId]
            if NumeroTipoID == 13:
                nombrecompleto = dfBalance[(dfBalance["Numero de identificacion"]==str(Id))]['Razón social'].iloc[0]
                if len(nombrecompleto.split())==1:
                    vectorPrimerApellido.append(nombrecompleto)
                    vectorSegundoApellido.append("")
                    vectorNombre.append("")
                    vectorOtrosNombres("")
                elif len(nombrecompleto.split())==2:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append("")
                    vectorOtrosNombres("")
                elif len(nombrecompleto.split())==3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())>3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append(nombrecompleto.split(maxsplit=3)[3])
            else:
                vectorPrimerApellido.append("")
                vectorSegundoApellido.append("")
                vectorNombre.append("")
                vectorOtrosNombres.append("")
            vectorCoincidencias.append(TiposDoc[tipoId])
        except Exception:
            vectorCoincidencias.append("NE")  
            vectorPrimerApellido.append("")
            vectorSegundoApellido.append("")
            vectorNombre.append("")
            vectorOtrosNombres.append("")
    dfBalance.insert(1,'Tipo de documento', vectorCoincidencias)
    dfBalance.insert(4,'Pais', 169)
    dfBalance.insert(3,'Otros nombres', vectorOtrosNombres)
    dfBalance.insert(3,'Primer nombre', vectorNombre)
    dfBalance.insert(3,'Segundo apellido', vectorSegundoApellido)
    dfBalance.insert(3,'Primer apellido', vectorPrimerApellido)
    return dfBalance

def separarCuentas(df):
    cuentas=df["Cuentas"]
    VectorCuentas = []
    VectorNits = []
    VectorNombres = []
    for cuenta in cuentas:  
        try:
            nit,name = cuenta.split(maxsplit=1)
        except Exception:
            try:
                cta = int(str(cuenta).strip().replace('.', ''))
            except Exception:
                cta = 0
            name = ""
            nit = ""
        VectorNombres.append(name)
        VectorCuentas.append(cta)
        VectorNits.append(nit)
    df.insert(1,"Razón social",VectorNombres)
    df.insert(1,"NIT",VectorNits)
    df.insert(1,"NumeroCuenta",VectorCuentas)
    return df

def ObtenerIngresos(df,ColumnaValorIngreso):
    ingresosAO = df[(df['NumeroCuenta']>413499)&(df['NumeroCuenta']<417000)&(df['NIT']!="")].iloc[:,ColumnaValorIngreso].sum()
    print('4135+55',': ',ingresosAO)#,df[(df['NumeroCuenta']==4135)].iloc[0]['             Descripción                                '].strip()
    devoluciones= df[(df['NumeroCuenta']<417600)&(df['NumeroCuenta']>417499)&(df['NIT']!="")].iloc[:,ColumnaValorIngreso].sum()
    print('4175',': ',devoluciones)#df[(df['NumeroCuenta']==4175)].iloc[0]['             Descripción                                '].strip(),
    ingresosO = df[(df['NumeroCuenta']<430000)&(df['NumeroCuenta']>420999)&(df['NIT']!="")].iloc[:,ColumnaValorIngreso].sum()
    print('42',': ',ingresosO)#df[(df['NumeroCuenta']==42)].iloc[0]['             Descripción                                '].strip(),
    print('TOTAL INGRESOS :',ingresosAO+devoluciones+ingresosO)
    return ingresosAO+devoluciones+ingresosO

def UnificarClientesPorCuenta(df,limiteInferiorCta,LimiteSuperiorCta,nombreDatoColumna,ColumnaValorIngreso):
    datosPorCliente = pd.DataFrame()
    listaSoloClientesIngreso = df[(df['NIT']!="")&(df['NumeroCuenta']>limiteInferiorCta)&(df['NumeroCuenta']<LimiteSuperiorCta)]
    listaUnica = listaSoloClientesIngreso["Razón social"].unique().tolist()
    VectorNombres = []
    VectorIdentificacion = []
    VectorTotales = []
    for cliente in listaUnica:
        total = listaSoloClientesIngreso[(listaSoloClientesIngreso['Razón social']==cliente)].iloc[:,ColumnaValorIngreso].sum()
        VectorIdentificacion.append(df[df['Razón social']==cliente]['NIT'].iloc[0])
        VectorNombres.append(cliente)
        VectorTotales.append(total)
    datosPorCliente.insert(0,"Razón social",VectorNombres)
    datosPorCliente.insert(0,"Numero de identificacion",VectorIdentificacion)
    datosPorCliente.insert(2,nombreDatoColumna,VectorTotales)
    return datosPorCliente

   
