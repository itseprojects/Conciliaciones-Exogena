# -*- coding: utf-8 -*-
"""
Created on Fri Oct 28 17:19:31 2022

@author: HaroldFerneyGomez
"""
from cmath import nan
from concurrent.futures import ThreadPoolExecutor
from itertools import repeat
import logging
from operator import index
import azure.functions as func
from datetime import datetime, timedelta, date
import pandas as pd
from azure.storage.blob import BlobSasPermissions, generate_blob_sas, BlobServiceClient
import urllib.parse
from math import ceil
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import unicodedata
import json
import pyodbc
import requests
# import grequests

# Datos del blob storage
account_name = 'itseblobdev' #'probepython'
account_key = 'd9sOh0WeqvVF66NQnyWKZWFL/KDje0LizX8UyFWpWX39lLX2C8fxnqRtYD2lOFvNp6aaayQsAq7T+AStvsHyew==' #Zas0npJX9ryEm4hmW/gatWr8aI91oOCvt+qbQKqWrZJCmhv5qh6S/w6ittYYaDBRjRnoxa0h+A8H+ASttcvrrQ=='
# container_name = 'data' #'test'

MupiosenBlob = "Municipios_de_Colombia.xlsx"
blob_name_to_save = 'Formato1001-'+ str(date.today())+'.xlsx' # Archivo excel a guardar en el blob correspondiente al formato

# para trabajar localmente, solicitado en el body del trigger equivalente a balanceFile
# FilePath = "balance 2021 Exeltis con terceros.xlsx"

# Definicion de los conceptos para el formato tipo SIESA
conceptos = {"5055":[515500,525500,725500], 
            "5056":[519520],
            "5002":[511000,521000,721000], # solo si la persona natural está en 2485: -20 hon no declarantes, o -21 declarantes, se reporta, sino en el f2276
            "5003":[530516],
            "5004":[513000,513500,514000,514500,515000,523500,523600,524000,524500,530505,722500,723000,723500,724000,724500],
            "5005":[512000,722000,169920],
            "5006":[530520,530521,530522,530523], # formato: se reportó en el 5063 excepto el 530523 intereses intercompany
            "5063":[531521], 
            "5007":[141000,143500,144000],
            "5008":[150800,152400,152500,152800,154000,158800],
            # "5010":[510565,510568,510571,520572,520575,520578],
            # "5011":[510574,510577,520568,520569],
            # "5012":[510580,520570],
            "5013":[],
            "5014":[], 
            "5015":[511505,511508,521505,521508,521535,530550,721500],
            "5066":[],
            "5058":[512500,522500],
            "5060":[],
            "5016":[161000,163000,164000,166000,517000,519525,519530,519595,519900,529510,529530,529540,529560,529595,530525,530526,530540,532010,539500,727000,729500],
            "5020":[],
            "5027":[],
            "5023":[],
            "5067":[],
            "5024":[],
            "5025":[],
            "5026":[],
            "5080":[],
            "5081":[],
            "5082":[],
            "5083":[],
            "5084":[],
            "5085":[],
            "5028":[],
            "5029":[],
            "5030":[],
            "5031":[],
            "5032":[169900,179900,189900,199900],
            "5033":[],
            "5034":[],
            "5035":[],
            "5019":[516500,726500],
            "5044":[],
            "5046":[],
            "5045":[],
            "5059":[],
            "5061":[],
            "5068":[],
            "5069":[],
            "5070":[],
            "5071":[],
            "5073":[],
            "5074":[],
            "5075":[],
            "5076":[],
            "5079":[],
            }

# Columnas adicionales
pagoNoDeducible = {"5016":[529521,530555,532005,539520,539581,539582]}
ivaColumnaO = [511535,511537,521570,521571]
# No se presenta retenciones para el id:800197268 => aduanas 
# va en cualquier concepto (el primero por default) que esté el tercero
reteFuentePracRenta = {"5002":[248520,248521],
                       "5004":[248540,248541,248542],
                       "5005":[248560,248561],
                       "5016":[248570,248571,248582,248584,248587,248595]}
reteIvaRegComun = [248710] 
reteIvaNoDomiciliada = [248711] 

# Función principal que permite analizar el balance, bd de usuarios y municipios de Colombia para formar el Formato 1001
    # requiere en el body la clave Balance y su valor es el nombre del balance en el Blob Storage
    # retorna cod 200 con el texto "Ejecución exitosa" si todo fue bien,
    # si hubo algun error retorna el texto "!! Ocurrió un error en la ejecución." más el posible origen del error
def main(req: func.HttpRequest) -> func.HttpResponse:
    Cliente = req.get_json().get('Cliente') # cliente1
    TipoBalance = req.get_json().get('TipoBalance') #"SIESA"
    blob_name_DB = req.get_json().get('BaseDeDatos') #'BASE DE DATOS EXELTIS.xls'
    balanceFile = req.get_json().get('Balance')
    idEjecucion = req.get_json().get('IdEjecucion')
    idProcedencia = req.get_json().get('IdProcedencia')
    planillas = req.get_json().get('Planillas')
    
    dfPlanillas = analizarPlanillas(planillas,Cliente)
    if TipoBalance=="SIESA":
        # Datos para la base de datos
        # blob_name_DB = 'BASE DE DATOS EXELTIS.xls'
        HeaderHojaDB = 0
        nombreHojaDB="Sheet1"
        exito = WorkSiesa(Cliente,balanceFile,blob_name_DB,idEjecucion,idProcedencia,dfPlanillas)
    else: 
        dicToReturn = {"error":"Tipo de balance no implementado"}
        exito = json.dumps(dicToReturn) 

    return func.HttpResponse(f"{exito}", status_code=200 )

def getData(urlPDF,cliente):
    url = "https://readpdfs.azurewebsites.net/api/readpdf"
    data = {
        "tipoPDF":"planilla",
        "id":3,
        "urlPDF":urlPDF,
        "passwordPDF":"",
        "cliente": cliente
    }
    r  = requests.post(url,json=data)
    return str(r.json()['ruta']).rsplit('/',1)[1]

def analizarPlanillas(planillas,Cliente):
    # Ingreso al blob storage
    blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
    DatosPlanillas = pd.DataFrame()
    with ThreadPoolExecutor() as executor:
        for planilla, ruta in zip(planillas, executor.map(getData,planillas,repeat(Cliente))):
            print(f'{planilla} resultado: {ruta}')
            blob_client = blob_service_client.get_blob_client(container = Cliente, blob = ruta.replace('%20',' '))
            downloader = blob_client.download_blob()
            Datos = pd.read_excel(downloader.readall(), sheet_name="Datos", header=0,engine='openpyxl')
            Datos =  Datos[~Datos['AFILIADOS'].isnull()]
            Datos =  Datos[~(Datos['RIESGO']=="TOTAL")]
            Datos = Datos.drop(['CODIGO','DV','AFILIADOS','INTERESES MORA','SALDOS E','VALOR A PAGAR'], axis=1)
            Datos['VALOR LIQUIDADO'] = Datos['VALOR LIQUIDADO'].apply(lambda x: int(x.replace(',','')))
            riesgos = Datos['RIESGO']
            conceptos = []
            concepto = ""
            for riesgo in riesgos:
                try:
                    if riesgo.__contains__('(ADMINISTRADORAS:'):
                        
                        concepto = "5012" if riesgo.__contains__('AFP') else "5011" if bool([x for x in ['ARL','EPS'] if(x in riesgo)]) else "5010"
                except:pass
                conceptos.append(concepto)
            Datos['Concepto'] = conceptos
            DatosPlanillas = pd.concat([DatosPlanillas, Datos], ignore_index=True ).groupby(['NIT','RIESGO','Concepto']).sum(numeric_only=True).reset_index()
    DatosPlanillas['VALOR LIQUIDADO'] = np.where(DatosPlanillas['Concepto']=="5012",75*DatosPlanillas['VALOR LIQUIDADO']/100,np.where(DatosPlanillas['Concepto']=="5011",75*DatosPlanillas['VALOR LIQUIDADO']/100,DatosPlanillas['VALOR LIQUIDADO']))
    DatosPlanillas['Pago o abono en cuenta no deducible'] = np.where(DatosPlanillas['Concepto']=="5012",25*DatosPlanillas['VALOR LIQUIDADO']/100,np.where(DatosPlanillas['Concepto']=="5011",25*DatosPlanillas['VALOR LIQUIDADO']/100,0))
    DatosPlanillas['NIT'] = DatosPlanillas['NIT'].apply(lambda x: int(x.replace(',','')))
    DatosPlanillas.rename(columns = {'RIESGO':'Razón social','VALOR LIQUIDADO':'Ingresos Brutos Recibidos','NIT':'Numero de identificacion'}, inplace = True)
    DatosPlanillas.insert(5,'NumeroCuenta',510500)
    return DatosPlanillas                                 

def WorkSiesa(container_name,balanceFile,blob_name_DB,idEjecucion,idProcedencia,dfPlanillas):  
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1" 
    try:
        # Ingreso al blob storage
        blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
        
        try: # Eliminación de archivo de salida si ya existe en el blob
            blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_to_save)
            blob_client.delete_blob()
        except:
            pass
        
        # leer balance y extraer columna del saldo total
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, nrows=0, header=10,engine='openpyxl')
        ColumnaValorIngreso = Datos.columns.get_loc("Saldo final a")
        
        # leer resto del balance
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=HeaderHojaBalance,engine='openpyxl')
        Datos = Datos[~Datos['Cuentas'].isnull()]
        DatosSeparados = separarCuentas(Datos)
        ColumnaValorIngreso += 3    # separaCuentas añade 3 columnas
        
        # Elaborar formato parte contable
        TercerosPorConcepto = GetClientsByConcept(DatosSeparados,ColumnaValorIngreso)
        TercerosPorConcepto = GetClientsByPagoNoDeducible(DatosSeparados,ColumnaValorIngreso,TercerosPorConcepto,'pagoNoDeducible')       
        # insertar planillas
        TercerosPorConcepto = pd.concat([TercerosPorConcepto,dfPlanillas], ignore_index=True ).reset_index()
        TercerosPorConcepto = GetIvaByClient(DatosSeparados, ColumnaValorIngreso,TercerosPorConcepto,'iva')
        TercerosPorConcepto.insert(6,'Iva mayor valor del costo o gasto no deducible', 0) # Siempre en 0
        TercerosPorConcepto = GetClientsByPagoNoDeducible(DatosSeparados,ColumnaValorIngreso,TercerosPorConcepto,'reteFuente') 
        TercerosPorConcepto.insert(8,'Retención en la fuente asumida en renta', 0)  # Siempre en 0
        TercerosPorConcepto = GetIvaByClient(DatosSeparados, ColumnaValorIngreso,TercerosPorConcepto,'reteIvaRegComun')
        TercerosPorConcepto = GetIvaByClient(DatosSeparados, ColumnaValorIngreso,TercerosPorConcepto,'reteIvaRegNoDomiciliada')
        
        # Realizar comprobaciones de la parte contable
        dfDiferencias = makeComprobations(TercerosPorConcepto,container_name,DatosSeparados,ColumnaValorIngreso)
        # Almacenar comprobaciones en la BD
        saveToBD(dfDiferencias,idEjecucion,idProcedencia)
        # Unir terceros por cada concepto
        TercerosPorConcepto = TercerosPorConcepto.groupby(['Concepto','Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()
        TercerosPorConcepto =TercerosPorConcepto .drop('index',axis=1)       
        # Leer base de datos de usuarios y municipios de Colombia
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_DB)
        downloader = blob_client.download_blob()
        dbUsers = pd.read_excel(downloader.readall(), sheet_name="Sheet1", header=0)#,engine='openpyxl')
        blob_client = blob_service_client.get_blob_client(container = "data", blob = MupiosenBlob)
        downloader = blob_client.download_blob()
        dbMupios = pd.read_excel(downloader.readall(), sheet_name="municipios", header=0,converters={'Código Municipio':str,'Código Departamento':str},engine='openpyxl')

        # Agregar datos de los terceros al formato
        TercerosPorConcepto = BuscarId(TercerosPorConcepto,dbUsers,dbMupios)
        
        # Limpieza y adecuación de datos y columnas
        TercerosPorConcepto["Pais"]= np.where(TercerosPorConcepto["Código depto"]=="","",TercerosPorConcepto["Pais"])
        TercerosPorConcepto["Razón social"]= np.where(TercerosPorConcepto["Primer apellido"]!="","",TercerosPorConcepto["Razón social"])
        TercerosPorConcepto = TercerosPorConcepto.drop(TercerosPorConcepto[(TercerosPorConcepto['Ingresos Brutos Recibidos']==0) & 
                                                                           (TercerosPorConcepto['Pago o abono en cuenta no deducible']==0) &
                                                                           (TercerosPorConcepto['Iva mayor valor del costo o gasto deducible']==0) &
                                                                           (TercerosPorConcepto['Iva mayor valor del costo o gasto no deducible']==0) &
                                                                           (TercerosPorConcepto['Retención en la fuente practicada en renta']==0) &
                                                                           (TercerosPorConcepto['Retención en la fuente asumida en renta']==0) &
                                                                           (TercerosPorConcepto['Retención en la fuente practicada IVA régimen común']==0) &
                                                                           (TercerosPorConcepto['Retención en la fuente practicada IVA no domiciliados']==0)].index)
        TercerosPorConcepto.rename(columns = {'Ingresos Brutos Recibidos':'Pago o abono en cuenta deducible'}, inplace = True)
        TercerosPorConcepto = TercerosPorConcepto.drop(TercerosPorConcepto[(TercerosPorConcepto['Numero de identificacion']=='Generico') & ((TercerosPorConcepto['Concepto']=='5007') | (TercerosPorConcepto['Concepto']=='5008'))].index) # eliminar registro para este #id
        TercerosPorConcepto['Pago o abono en cuenta deducible'] = TercerosPorConcepto['Pago o abono en cuenta deducible'].apply(np.ceil)
        TercerosPorConcepto['Pago o abono en cuenta no deducible'] = TercerosPorConcepto['Pago o abono en cuenta no deducible'].apply(np.ceil)
        TercerosPorConcepto['Iva mayor valor del costo o gasto deducible'] = TercerosPorConcepto['Iva mayor valor del costo o gasto deducible'].apply(np.ceil)
        TercerosPorConcepto['Iva mayor valor del costo o gasto no deducible'] = TercerosPorConcepto['Iva mayor valor del costo o gasto no deducible'].apply(np.ceil)
        TercerosPorConcepto['Retención en la fuente practicada en renta'] = TercerosPorConcepto['Retención en la fuente practicada en renta'].apply(np.ceil)
        TercerosPorConcepto['Retención en la fuente asumida en renta'] = TercerosPorConcepto['Retención en la fuente asumida en renta'].apply(np.ceil)
        TercerosPorConcepto['Retención en la fuente practicada IVA régimen común'] = TercerosPorConcepto['Retención en la fuente practicada IVA régimen común'].apply(np.ceil)
        TercerosPorConcepto['Retención en la fuente practicada IVA no domiciliados'] = TercerosPorConcepto['Retención en la fuente practicada IVA no domiciliados'].apply(np.ceil)
        TercerosPorConcepto = TercerosPorConcepto.drop(['NumeroCuenta'], axis=1) # Eliminar columna no necesaria
        
        # Ajustar parte estética y almacenar en el BLob Storage
        PutColorsAnsSaveToBlob(TercerosPorConcepto,container_name)
            

        dicToReturn = {
            "error":"ninguno",
            "ruta":f'https://{account_name}.blob.core.windows.net/{container_name}/{blob_name_to_save}'
            }

    except Exception as e:
        dicToReturn = {"error":f"{e}"}
    return json.dumps(dicToReturn)

def saveToBD(comprobations,idEjecucion,idProcedencia):
    server = 'rbhaxa.database.windows.net' 
    database = 'haxa' 
    username = 'rbitse' 
    password = 'QbLnBh29XUrDpzX'
    driver= '{ODBC Driver 17 for SQL Server}'# {SQL Server}
    cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server}'+';DATABASE='+database+';ENCRYPT=yes;UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    for table_name in cursor.tables(tableType='TABLE'):
        print(table_name)
        
    for i in range(len(comprobations)): 
        nameDiff = comprobations.iloc[i,1].strip()#[0:35]
        insert_stmt = (
                        "INSERT INTO Diferencias (id_ejecuccion, id_procedencia, nombre_diferencia ,comprobacion, numeroc,observaciones,valor_HAXA) \
                        VALUES (?,?,?,?,?,?,?)"
                        )
        data = (idEjecucion,idProcedencia,comprobations.iloc[i,1].strip(),comprobations.iloc[i,2],str(comprobations.iloc[i,0]),comprobations.iloc[i,4],comprobations.iloc[i,3])
        # insertar registro en bd
        cursor.execute(insert_stmt, data)
        cnxn.commit()
    return None

def makeComprobations(TercerosPorConcepto,container_name,DatosSeparados,ColumnaValorIngreso):
    tablaParaComprobar = ExtraerCtas(DatosSeparados)
    tablaParaComprobar.columns.values[2] = "ValorContable"
    tablaParaComprobar['Formato'] = 0
    tablaParaComprobar['Observaciones'] = ""
    
    for concepto in conceptos:
        for clave in conceptos[concepto]:
            clave4Digitos = int(str(clave)[0:4])
            valorEnFormato = TercerosPorConcepto[(TercerosPorConcepto['Concepto']==concepto) & (TercerosPorConcepto['NumeroCuenta']==clave)].loc[:,'Ingresos Brutos Recibidos'].sum()
            tablaParaComprobar["Formato"]+= np.where(tablaParaComprobar["NumeroCuenta"]==clave4Digitos,valorEnFormato,0)#tablaParaComprobar["Formato"])
            try:
                ind = tablaParaComprobar[(tablaParaComprobar["NumeroCuenta"]==clave4Digitos)].index
                pasaAjustePeso = True if abs(tablaParaComprobar.loc[ind,'ValorContable'].iloc[0]-tablaParaComprobar.loc[ind,'Formato'].iloc[0])<1000 else False
                tablaParaComprobar.loc[ind,'Observaciones']='' if pasaAjustePeso else "Valor reportado es diferente al valor contable"
                if clave4Digitos == 5195:
                    print(tablaParaComprobar.loc[ind,'ValorContable'].iloc[0],' : contable - formato: ',valorEnFormato)
            except Exception as e:
                pass
    #tablaParaComprobar["Observaciones"]= np.where(tablaParaComprobar["ValorContable"]==tablaParaComprobar["Formato"],tablaParaComprobar["Observaciones"],"No Coincide")
    return tablaParaComprobar

# Función que agrega color según el tipo de contenido de las columnas y guarda archivo excel en el Blob Storage,
#   Asigna el color rojo para advertir sobre datos relacionados de la base de datos de usuarios y amarillo para datos relacionados con el balance
#   requiere de: 
#       Datos = df del formato con todas las columnas
#       account_name, account_key, blob_name_to_save
#   retorna: none
def PutColorsAnsSaveToBlob(Datos,container_name):
    fillOrange = PatternFill(patternType='solid', fgColor='FCBA03')
    fillRed = PatternFill(patternType='solid', fgColor='EE1111')
    # FilePath = "balance 2021 Exeltis con terceros.xlsx"
    nombreHojaBalance= "iva+Names" # Nombre de la hoja a guardar
    blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
    # blob_client = blob_service_client.get_blob_client(container = container_name, blob = FilePath)
    # downloader = blob_client.download_blob()
    # Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=0)
    buffer = BytesIO()
    excel_buf = Datos.to_excel(sheet_name=nombreHojaBalance,excel_writer=buffer,index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(Datos, index=False, header=True):
        ws.append(r)
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    table = openpyxl.worksheet.table.Table(ref='A1:T'+str(ws.max_row), displayName='Formato1001ByHG', tableStyleInfo=mediumStyle)
    ws.add_table(table)
    
    # Agrega color a las celdas segun las siguientes normas
    for row, rowval in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in rowval:
            
            # rojo si Tipo de doc = NE, no encontrado en bd usuarios
            if f"{cell.column_letter}"=="B" and f"{cell.value}"=="NE":
                ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
                
            # rojo si Direccion, departamente, municipio o país está vacio
            elif f"{cell.column_letter}"=="I" or f"{cell.column_letter}"=="J" or f"{cell.column_letter}"=="K" or f"{cell.column_letter}"=="L":
                if f"{cell.value}"=="" or f"{cell.value}"=="None" or f"{cell.value}"=="nan":
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
                    
            # Naranja si pagos o abonos en cuenta menores a 100 mil pesos y diferentes de cero
            elif f"{cell.column_letter}"=="M" or f"{cell.column_letter}"=="N": 
                if cell.value<=100 and cell.value!=0:
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillOrange
                    
            # Naranja si Ivas o Retenciones son negativas
            elif (   f"{cell.column_letter}"=="O" or f"{cell.column_letter}"=="P" or f"{cell.column_letter}"=="Q"
                  or f"{cell.column_letter}"=="R" or f"{cell.column_letter}"=="S" or f"{cell.column_letter}"=="T"): 
                if cell.value<0 :
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillOrange
                    
    # Subir excel al blob Storage
    abs_container_client = blob_service_client.get_container_client(container=container_name)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        blockBlob = abs_container_client.upload_blob(name=blob_name_to_save,data=stream)
    return None

# Función para obtener el iva, la retención IVA régimen común o la retención IVA no domicialiados segun el modo indicado
#   requiere de: 
#       DatosSeparados = balance con cta separada de los datos del tercero 
#       ColumnaValorIngreso = Columna del valor del ingreso
#       TercerosPorConcepto = df de los terceros hallados por cada concepto
#       cualColumna = modo de operación de la función: "iva", "reteIvaRegComun" u "otro valor para no domiciliados"
#   retorna df con columnas similares a TercerosPorIva más 1 columna extra según el modo de operación
def GetIvaByClient(DatosSeparados, ColumnaValorIngreso,TercerosPorConcepto, cualColumna):
    TercerosPorIva = pd.DataFrame()
    TercerosPorIva['Concepto'] = None
    TercerosPorIva['Numero de identificacion']= None
    TercerosPorIva['Razón social'] = None
    TercerosPorIva['Ingresos Brutos Recibidos'] = None
    TercerosPorIvaParaComparar = TercerosPorIva
    TercerosPorIvaParaComparar['Iva mayor valor del costo o gasto deducible' if cualColumna=="iva" else 'Retención en la fuente practicada IVA régimen común' if 
                               cualColumna=='reteIvaRegComun' else 'Retención en la fuente practicada IVA no domiciliados'] = None
    # recorrer vector indicado segun el modo de opereción
    for ivaCount in ivaColumnaO if cualColumna=="iva" else reteIvaRegComun if cualColumna=='reteIvaRegComun' else reteIvaNoDomiciliada:
        # print(">> "+str(ivaCount))
        data = UnificarClientesPorCuenta(DatosSeparados,ivaCount,ivaCount+1,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
        # print(data.columns)
        if ~(data.empty) and cualColumna!='iva':
            data = data.drop(data[(data['Numero de identificacion']=='800197268')].index) # eliminar registro para este #id DIRECCION DE IMPUESTOS Y ADUANAS NACIONALES
        TercerosPorIva = pd.concat([TercerosPorIva, data], ignore_index=True ).groupby(['Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()
            
    # Buscar en TercerosPorConcepto si existen los clientes en cualquier concepto y se fusionan los datos
    if ~TercerosPorIva.empty: 
        for ind in TercerosPorIva.index :
            try:
                firstElement = TercerosPorConcepto[TercerosPorConcepto['Numero de identificacion']==TercerosPorIva['Numero de identificacion'][ind]].groupby('Numero de identificacion').first()
            except Exception:
                firstElement = TercerosPorIva[ind]
                firstElement['Ingresos Brutos Recibidos']=0
            firstElement.insert(3,'Iva mayor valor del costo o gasto deducible'if cualColumna=="iva" else 'Retención en la fuente practicada IVA régimen común' if 
                               cualColumna=='reteIvaRegComun' else 'Retención en la fuente practicada IVA no domiciliados', (1 if cualColumna=="iva" else -1)*TercerosPorIva['Ingresos Brutos Recibidos'][ind])
            TercerosPorIvaParaComparar = pd.concat([firstElement.reset_index() ,TercerosPorIvaParaComparar.loc[:]])#.reset_index(drop=True)
            # TercerosPorIvaParaComparar.append(firstElement,ignore_index=True)
        TercerosPorConcepto = pd.merge(TercerosPorConcepto, TercerosPorIvaParaComparar,  how='outer', indicator=False).fillna(value=0)
        # print(TercerosPorConcepto)
    # print(TercerosPorConcepto[['Razón social','Ingresos Brutos Recibidos']])
    return TercerosPorConcepto   

# Función que obtiene los terceros a partir del diccionario conceptos
#   requiere de: 
#       DatosSeparados = balance con cta separada de los datos del tercero 
#       ColumnaValorIngreso = Columna del valor del ingreso
#   retorna un df con columnas equivalentes a las de TercerosPorConcepto
def GetClientsByConcept(DatosSeparados, ColumnaValorIngreso):
    TercerosPorConcepto = pd.DataFrame()
    TercerosPorConcepto['Concepto'] = None
    TercerosPorConcepto['Numero de identificacion']= None
    TercerosPorConcepto['Razón social'] = None
    TercerosPorConcepto['Ingresos Brutos Recibidos'] = None
    # print(TercerosPorConcepto.columns)
    for clave in conceptos:
        estadoiteracion=True    # variable que permite obtener datos de las cuentas que inician por 15 y 16 una sola vez 
                                    # si existe la cta 1588 o 1688 respectivamente 
        soloConcepto = pd.DataFrame()
        soloConcepto['Numero de identificacion']= None
        soloConcepto['Razón social'] = None
        soloConcepto['Ingresos Brutos Recibidos'] = None
        #soloConcepto['NumeroCuenta'] = 0
        # print(soloConcepto.columns)
        for conc in conceptos[clave]:
            # print(">> "+str(conc))
            if str(conc).startswith("15"):
                data1 = UnificarClientesPorCuenta(DatosSeparados,158800,159000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                if data1.empty: 
                    data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ 1000),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                elif estadoiteracion:
                    estadoiteracion=False
                    data=data1
                else: data=pd.DataFrame()
            elif str(conc).startswith("16"):
                if str(conc).startswith("1699"):
                    data = UnificarClientesPorCuenta(DatosSeparados,168800,169000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                else:
                    data1 = UnificarClientesPorCuenta(DatosSeparados,168800,169000,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                    if data1.empty: 
                        data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ 1000),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                    elif estadoiteracion:
                        estadoiteracion=False
                        data=data1
                    else: data=pd.DataFrame()
            else: data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ (100 if str(conc).endswith("00") else 1)),"Ingresos Brutos Recibidos",ColumnaValorIngreso)

            data.insert(0,'NumeroCuenta', conc)
            # soloConcepto00 = soloConcepto00.combine_first(data)
            
            
            # uncomment **********
            # Agrupa los terceros por su #id y razon social y suma sus totales
            soloConcepto = pd.concat([soloConcepto, data], ignore_index=True )#.groupby(['Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()
            # Agrega todos los terceros únicos en el concepto al df de salida
        # solo si la persona natural está en 2485: -20 hon no declarantes, o -21 declarantes, se reporta, sino en el f2276
        if clave=='5002':
            data0 = UnificarClientesPorCuenta(DatosSeparados,248520,248522,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
            mascara = soloConcepto["Numero de identificacion"].isin(data0["Numero de identificacion"])
            soloConcepto = soloConcepto[mascara]
        if ~soloConcepto.empty: 
            soloConcepto.insert(0,'Concepto', clave)
            TercerosPorConcepto = pd.concat([TercerosPorConcepto, soloConcepto],ignore_index=True)
            # print(soloConcepto)
    # print(TercerosPorConcepto[['Razón social','Ingresos Brutos Recibidos']])
    return TercerosPorConcepto       

# Función que obtiene los terceros a partir del diccionario pagoNoDeducible o reteFuentePractRenta
#   requiere de: 
#       DatosSeparados = balance con cta separada de los datos del tercero 
#       ColumnaValorIngreso = Columna del valor del ingreso
#       Terceros = df de los terceros hallados por cada concepto
#       modo = 2 modo de operación de la función, "pagoNoDeducible" o "otro"
#   retorna df con columnas equivalentes a las de Terceros más 1 columna extra según el modo de operación
def GetClientsByPagoNoDeducible(DatosSeparados, ColumnaValorIngreso,Terceros, modo):
    TercerosPorConcepto = pd.DataFrame()
    TercerosPorConcepto['Concepto'] = None
    TercerosPorConcepto['Numero de identificacion']= None
    TercerosPorConcepto['Razón social'] = None
    TercerosPorConcepto['Ingresos Brutos Recibidos'] = None
    TercerosParaComparar = TercerosPorConcepto
    TercerosParaComparar['Pago o abono en cuenta no deducible' if modo=="pagoNoDeducible" else 'Retención en la fuente practicada en renta'] = None
    
    # print(TercerosPorConcepto.columns)
    for clave in pagoNoDeducible if modo=="pagoNoDeducible" else reteFuentePracRenta:
        soloConcepto = pd.DataFrame()
        soloConcepto['Numero de identificacion']= None
        soloConcepto['Razón social'] = None
        soloConcepto['Ingresos Brutos Recibidos'] = None
        for conc in pagoNoDeducible[clave] if modo=="pagoNoDeducible" else reteFuentePracRenta[clave]:
            data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ 1),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
            data.insert(0,'NumeroCuenta', conc) # ********************
            if  modo!="pagoNoDeducible": # modo obtener retencion
                data = data.drop(data[(data['Numero de identificacion']=='800197268')].index) # eliminar registro para este #id
            soloConcepto = pd.concat([soloConcepto, data], ignore_index=True ) # uncomment.groupby(['Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()    
        if ~soloConcepto.empty: 
            for ind in soloConcepto.index :
                try:    # buscar si ya existe en Terceros ese cliente para ese concepto, de lo contrario lo agrega con ingresos = 0
                    firstElement = Terceros[(Terceros['Concepto']==clave)&(Terceros['Numero de identificacion']==soloConcepto['Numero de identificacion'][ind])].groupby('Numero de identificacion').first()
                    if firstElement.empty: 
                        firstElement = soloConcepto.loc[ind:ind]
                        firstElement['Ingresos Brutos Recibidos']=0
                        firstElement['Concepto']=clave
                        #firstElement['NumeroCuenta'] = conc
                except Exception:
                    firstElement = soloConcepto.loc[ind:ind]
                    firstElement['Ingresos Brutos Recibidos']=0
                    firstElement['Concepto']=clave
                
                # Agregar columnas según el modo de operación de la función
                firstElement.insert(3,'Pago o abono en cuenta no deducible' if modo=="pagoNoDeducible" else 'Retención en la fuente practicada en renta', 
                                    (1 if modo=="pagoNoDeducible" else -1)*soloConcepto['Ingresos Brutos Recibidos'][ind])
                firstElement = firstElement.reset_index()
                
                # Agregar datos encontrados por el concepto a TercerosParaComparar
                TercerosParaComparar = pd.concat([firstElement ,TercerosParaComparar.loc[:]])#.reset_index(drop=True)
                
                # seleccionar columnas necesarias
                TercerosParaComparar = TercerosParaComparar[["Numero de identificacion","Razón social","Ingresos Brutos Recibidos",
                                                         "Pago o abono en cuenta no deducible" if modo=="pagoNoDeducible" else 'Retención en la fuente practicada en renta',"Concepto",'NumeroCuenta']]
            # TercerosPorIvaParaComparar.append(firstElement,ignore_index=True)
    # Fusionar con el df de salida agregando las columnas segun el modo de operación de la función
    TercerosPorConcepto = pd.merge(Terceros, TercerosParaComparar,  how='outer', on=['Numero de identificacion','Concepto','Ingresos Brutos Recibidos','Razón social']).fillna(value=0)

    TercerosPorConcepto['NumeroCuenta'] = np.where(TercerosPorConcepto['NumeroCuenta_x']>0,TercerosPorConcepto["NumeroCuenta_x"],TercerosPorConcepto['NumeroCuenta_y']) 
    TercerosPorConcepto = TercerosPorConcepto.drop(['NumeroCuenta_x', 'NumeroCuenta_y'], axis=1)
    return TercerosPorConcepto  

# Función que lee la columna Cuentas, separa nits de razón social y las cuentas
#   requiere de:
#       df = df del balance
#   retorna: df más 3 columnas extras: Razón social, NIT y NumeroCuenta
def separarCuentas(df):
    cuentas=df["Cuentas"]
    if "Descripción" in cuentas[0]:
        cuentas=cuentas[1:]
        df = df[1:]
    VectorCuentas = []
    VectorNits = []
    VectorNombres = []
    for cuenta in cuentas:  
        try:
            nit,name = cuenta.split(maxsplit=1)
        except Exception:
            try:
                cta = int(str(cuenta).strip().replace('.', ''))
            except Exception:
                cta = 0
            name = ""
            nit = ""
        VectorNombres.append(name)
        VectorCuentas.append(cta)
        VectorNits.append(nit)
    df.insert(1,"Razón social",VectorNombres)
    df.insert(1,"NIT",VectorNits)
    df.insert(1,"NumeroCuenta",VectorCuentas)
    return df

# Función para el formato 1007 que obtiene el valor total de:
#   ingresos: ctas 4135 y 4155,
#   devoluciones: cta 4175,
#   otros ingresos: ctas 42 y
# retorna la suma de de los anteriores (ingresos+devoluciones)
def ObtenerIngresos(df,ColumnaValorIngreso):
    ingresosAO = df[(df['NumeroCuenta']>413499)&(df['NumeroCuenta']<417000)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('4135+55',': ',ingresosAO)#,df[(df['NumeroCuenta']==4135)].iloc[0]['             Descripción                                '].strip()
    devoluciones= df[(df['NumeroCuenta']<417600)&(df['NumeroCuenta']>417499)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('4175',': ',devoluciones)#df[(df['NumeroCuenta']==4175)].iloc[0]['             Descripción                                '].strip(),
    ingresosO = df[(df['NumeroCuenta']<430000)&(df['NumeroCuenta']>420999)&(df['NIT']!="")].iloc[:, ColumnaValorIngreso].sum()
    print('42',': ',ingresosO)#df[(df['NumeroCuenta']==42)].iloc[0]['             Descripción                                '].strip(),
    print('TOTAL INGRESOS :',ingresosAO+devoluciones+ingresosO)
    return ingresosAO+devoluciones+ingresosO

# Función que obtiene el valor de la columnaValorIngreso para terceros únicos entre un rango definido,
#   requiere de: 
#       df = df del balance 
#       limiteInferiorCta = rango mínimo del valor de cta a buscar, incluido
#       LimiteSuperiorCta = rango máximo del valor de la busqueda, no incluido 
#       nombreDatoColumna = nombre de la columna a insertar en el df de salida
#       ColumnaValorIngreso = valor numérico de la columna Saldo final a (del balance)
#   retorna df con 3 columnas: Razón social, Numero de identificacion y valor de nombreDatoColumna 
def UnificarClientesPorCuenta(df,limiteInferiorCta,LimiteSuperiorCta,nombreDatoColumna,ColumnaValorIngreso):
    datosPorCliente = pd.DataFrame()
    listaSoloClientesIngreso = df[(df['NIT']!="")&(df['NumeroCuenta']>=limiteInferiorCta)&(df['NumeroCuenta']<LimiteSuperiorCta)]
    listaUnica = listaSoloClientesIngreso["Razón social"].unique().tolist()
    VectorNombres = []
    VectorIdentificacion = []
    VectorTotales = []
    for cliente in listaUnica:
        total = listaSoloClientesIngreso[(listaSoloClientesIngreso['Razón social']==cliente)].iloc[:, ColumnaValorIngreso].sum()
        # print(cliente,total)
        VectorIdentificacion.append(df[df['Razón social']==cliente]['NIT'].iloc[0])
        VectorNombres.append(cliente)
        VectorTotales.append(total)
        
    datosPorCliente.insert(0,"Razón social",VectorNombres)
    datosPorCliente.insert(0,"Numero de identificacion",VectorIdentificacion)
    datosPorCliente.insert(2,nombreDatoColumna,VectorTotales)
    # print(limiteInferiorCta," hasta limite ",LimiteSuperiorCta,datosPorCliente)
    return datosPorCliente

# Función que guarda archivo excel localmente mediante 2 métodos: pd.to_excel y openpyxl.save (este permite agregar color a las celdas)
#   Método pandas requiere de:
#       df = df a guardar como excel
#       nombreHoja = nombre de la hoja a guardar
#   Método opnpyxl requiere de:
#       FilePath = ruta de archivo excel a poner color y guardar con otro nombre definido
#   retorna: null
# def GuardarExcel(df, nombreHoja):
    # ExcelWorkbook = load_workbook(FilePath)
    # writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
    # writer.book = ExcelWorkbook
    # df.to_excel(writer, sheet_name = nombreHoja ,index=False)
    # writer.save()
    # writer.close()
    # wb = openpyxl.load_workbook(FilePath)
    # ws = wb[nombreHoja]
    # fillOrange = PatternFill(patternType='solid', fgColor='FCBA03')
    # fillRed = PatternFill(patternType='solid', fgColor='EE1111')
    # for row, rowval in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    #     for cell in rowval:
    #         if f"{cell.column_letter}"=="B" and f"{cell.value}"=="NE":
    #             ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
    #         elif f"{cell.column_letter}"=="I" or f"{cell.column_letter}"=="J" or f"{cell.column_letter}"=="K" or f"{cell.column_letter}"=="L":
    #             if f"{cell.value}"=="" or f"{cell.value}"=="None":
    #                 ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
    #         elif f"{cell.column_letter}"=="M" or f"{cell.column_letter}"=="N" or f"{cell.column_letter}"=="O":
    #             if cell.value<=100 and cell.value!=0:
    #                 ws[f"{cell.column_letter}{cell.row}"].fill = fillOrange
    # wb.save("balance 2021 Exeltis con terceros1.xlsx")
    # wbbb = pd.DataFrame(wb.values)

# Función que busca el tercero en la bd según su #id y también le agrega su Mpio 
#   requiere de:
#       dfBalance = df de terceros por cada concepto
#       db = df de la base de datos de los usuarios
#       dbMupios = df de la bd de los municipios de Colombia
#   retorna: dfBalance más columnas con información del tercero
def BuscarId(dfBalance,bd,dbMupios):
    TiposDoc={'C':13,'E':31,'N':31,'O':43,'X':31}   # bd asigna estas letras según el tipo de doc, se asignan # según normativa del formato
    ListadoIds = dfBalance["Numero de identificacion"].tolist()
    vectorDireccion = []
    vectorCodMpio = []
    vectorCodDepto = []
    vectorCoincidencias = []
    vectorPrimerApellido = []
    vectorSegundoApellido = []
    vectorNombre = []
    vectorOtrosNombres = []
    bd['Código']=bd['Código'].apply(lambda x: x.strip())
    for Id in ListadoIds:
        # Buscar si existe el tercero en la db y agrega su información, si no existe agrega NE a columna tipo de documento
        try:
            tipoId = bd[(bd['Código']==str(Id))]['Tipo de identificación'].iloc[0]
            NumeroTipoID = TiposDoc[tipoId]
            # si es una Cédula, separa por nombres y apellidos. De lo contrario agrega texto vacio
            if NumeroTipoID == 13:
                nombrecompleto = dfBalance[(dfBalance["Numero de identificacion"]==str(Id))]['Razón social'].iloc[0]
                if len(nombrecompleto.split())==1:
                    vectorPrimerApellido.append(nombrecompleto)
                    vectorSegundoApellido.append("")
                    vectorNombre.append("")
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())==2:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append("")
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())==3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())>3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0]+" "+nombrecompleto.split()[1] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[0]+" "+nombrecompleto.split()[1]+" "+nombrecompleto.split()[2] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[2] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[3] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[3] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[4] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[2])
                    vectorOtrosNombres.append(nombrecompleto.split(maxsplit=4)[4] if nombrecompleto.split()[0].lower()=="del"
                                            else "" if len(nombrecompleto.split())==5 and nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la"
                                            else nombrecompleto.split(maxsplit=5)[5] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" and len(nombrecompleto.split())>5
                                            else nombrecompleto.split(maxsplit=3)[3])
            else:
                vectorPrimerApellido.append("")
                vectorSegundoApellido.append("")
                vectorNombre.append("")
                vectorOtrosNombres.append("")
            vectorCoincidencias.append(TiposDoc[tipoId])
        except Exception:
            vectorCoincidencias.append("NE")  
            vectorPrimerApellido.append("")
            vectorSegundoApellido.append("")
            vectorNombre.append("")
            vectorOtrosNombres.append("")
        
        # Buscar la ciudad en dbMupios, uso de unicodedata para cambiar acentos
        try:
            mpio = bd[(bd["Código"]==str(Id))]['Ciudad'].iloc[0]
            mpio = unicodedata.normalize('NFKD', mpio).encode('ASCII', 'ignore').decode("UTF-8")
            dbMupios['Nombre Municipio'] = [unicodedata.normalize('NFKD', x).encode('ASCII', 'ignore').decode("UTF-8") for x in dbMupios['Nombre Municipio']]
            firstMpio = dbMupios.loc[dbMupios['Nombre Municipio'].str.contains(mpio,case=False)].sort_values('Nombre Municipio')
            
            # Si no encuentra la ciudad o el tipo de identificación es X para el tercero, probablemente no sea de Colombia
            if firstMpio.empty or bd[(bd["Código"]==str(Id))]['Tipo de identificación'].iloc[0]=="X":
                firstMpio=pd.Series({'Código Departamento':'', 'Código Municipio':"", 'Nombre Departamento':"", 'Nombre Municipio':"", 'Tipo: Municipio / Isla / Área no municipalizada':""}, )
            # Selecciona el primer municipio de firstMpio
            else: firstMpio = firstMpio.iloc[0,:]
            vectorDireccion.append(bd[(bd["Código"]==str(Id))]['Dirección 1'].iloc[0]) 
            vectorCodDepto.append(firstMpio['Código Departamento'])
            vectorCodMpio.append(firstMpio['Código Municipio'])
        except Exception:
            vectorDireccion.append("") 
            vectorCodDepto.append("")
            vectorCodMpio.append("")
    dfBalance.insert(1,'Tipo de documento', vectorCoincidencias)
    dfBalance.insert(4,'Pais', 169)
    dfBalance.insert(3,'Otros nombres', vectorOtrosNombres)
    dfBalance.insert(3,'Primer nombre', vectorNombre)
    dfBalance.insert(3,'Segundo apellido', vectorSegundoApellido)
    dfBalance.insert(3,'Primer apellido', vectorPrimerApellido)
    dfBalance.insert(8,'Código mpc', vectorCodMpio)
    dfBalance.insert(8,'Código depto', vectorCodDepto)
    dfBalance.insert(8,'Dirección', vectorDireccion)
    return dfBalance

# Función local que extrae las ctas del balance sin los terceros, no usada 
#   requiere de:
#       FilePath, nombreHojaBalance, HeaderHojaBalance
#   retorna: dfBalance solo las 2 columnas iniciales
# uso:
#       df = ExtraerCtas()
#       GuardarExcel(df, 'soloCtas')
def ExtraerCtas(Datos):
    # HeaderHojaBalance=11
    # nombreHojaBalance= "Hoja 1"
    
    try:
        # Datos = pd.read_excel(FilePath, sheet_name=nombreHojaBalance, header=HeaderHojaBalance)
        Datos =  Datos[Datos['NumeroCuenta'].apply( lambda x: str(x).startswith(('14','15','5','6','7')) and x<10000)]
        Datos = Datos.iloc[:,[1,4,8]]
        print(Datos)
        return Datos
    except : False


# extras:
# Api para buscar municipios de colombia
# No funciona si no se dá el nombre correcto, ej: Medellín no encontrado sin tilde
# Mpio = "Cali"
# headers = {
#         'Content-Type': 'application/json'
#     }
# doc API Mpios Colombia https://dev.socrata.com/foundry/www.datos.gov.co/xdk5-pm3f
# response = requests.request("GET", "https://www.datos.gov.co/resource/xdk5-pm3f.json?municipio="+Mpio, headers=headers)
# print(response.text)

