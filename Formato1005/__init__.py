# -*- coding: utf-8 -*-
"""
Created on Fri Apr 13 19:17:31 2022

@author: AndresFelipeMarciales
"""

import azure.functions as func
from datetime import date
import pandas as pd
from math import ceil
from azure.storage.blob import BlobServiceClient
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import json
import pyodbc

# Datos del blob storage
account_name = 'itseblobdev' #'probepython'
account_key = 'd9sOh0WeqvVF66NQnyWKZWFL/KDje0LizX8UyFWpWX39lLX2C8fxnqRtYD2lOFvNp6aaayQsAq7T+AStvsHyew==' #Zas0npJX9ryEm4hmW/gatWr8aI91oOCvt+qbQKqWrZJCmhv5qh6S/w6ittYYaDBRjRnoxa0h+A8H+ASttcvrrQ=='
# container_name = 'data' #'test'

MupiosenBlob = "Municipios_de_Colombia.xlsx"
blob_name_to_save = 'Formato1005-'+ str(date.today())+'.xlsx' # Archivo excel a guardar en el blob correspondiente al formato

# Función principal que permite analizar el balance, bd de usuarios y municipios de Colombia para formar el Formato 1001
    # requiere en el body la clave Balance y su valor es el nombre del balance en el Blob Storage
    # retorna cod 200 con el texto "Ejecución exitosa" si todo fue bien,
    # si hubo algun error retorna el texto "!! Ocurrió un error en la ejecución." más el posible origen del error
def main(req: func.HttpRequest) -> func.HttpResponse:
    Cliente = req.get_json().get('Cliente') # cliente1
    TipoBalance = req.get_json().get('TipoBalance') #"SIESA"
    blob_name_DB = req.get_json().get('BaseDeDatos') #'BASE DE DATOS EXELTIS.xls'
    balanceFile = req.get_json().get('Balance')
    idEjecucion = req.get_json().get('IdEjecucion')
    idProcedencia = req.get_json().get('IdProcedencia')
    
    if TipoBalance=="SIESA":
        # Datos para la base de datos
        # blob_name_DB = 'BASE DE DATOS EXELTIS.xls'
        HeaderHojaDB = 0
        nombreHojaDB="Sheet1"
        exito = WorkSiesa(Cliente,balanceFile,blob_name_DB,idEjecucion,idProcedencia)
    else: 
        dicToReturn = {"error":"Tipo de balance no implementado"}
        exito = json.dumps(dicToReturn)

    return func.HttpResponse(f"{exito}", status_code=200 )

def WorkSiesa(container_name,balanceFile,blob_name_DB,idEjecucion,idProcedencia):
    """
    Elabora y guarda el formato 1005 en el Blob Storage, almacena las comprobaciones en BD.
    Args:
        container_name (str): Nombre del contenedor - cliente.
        balanceFile (str): Nombre del blob para el balance.
        blob_name_DB (str): Nombre del blob para la BD de Terceros.
        idEjecucion (int): Número del registro de la ejecución.
        idProcedencia (int): Numero del registro de la procedencia.
    Returns:
        dicToReturn: json con información de la ejecución de la función.
    """ 
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1" 
    try:
        # Ingreso al blob storage
        blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
        
        try: # Eliminación de archivo de salida si ya existe en el blob
            blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_to_save)
            blob_client.delete_blob()
        except:
            pass
        
        # leer balance y extraer columna del saldo total
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, nrows=0, header=10,engine='openpyxl')
        ColumnaValorIngreso = Datos.columns.get_loc("Saldo final a")
        
        # leer resto del balance
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=HeaderHojaBalance,engine='openpyxl')
        Datos = Datos[~Datos['Cuentas'].isnull()]
        DatosSeparados = separarCuentas(Datos)
        ColumnaValorIngreso += 3    # separaCuentas añade 3 columnas
        
        Totales = ObtenerImpuesto(DatosSeparados,ColumnaValorIngreso)

        # Agrupar clientes de clientas de interes
        ImpuestoClientes = UnificarClientesPorCuenta(DatosSeparados,240832,240850,"Impuesto Descontable",ColumnaValorIngreso)
        ImpuestoDescontable = ImpuestoClientes.sort_values(by = ['Impuesto Descontable'], ascending = [True], na_position = 'last',ignore_index=True)

        ImpuestoClientesDev = UnificarClientesPorCuenta(DatosSeparados,240820,240820,"IVA resultante devoluciones",ColumnaValorIngreso)
        ImpuestosDevolucion = ImpuestoClientesDev.sort_values(by = ['IVA resultante devoluciones'], ascending = [True], na_position = 'last',ignore_index=True)

        #Unificar Impuesto descontable e IVA resultante devoluciones
        ImpuestoUnificado = pd.merge(ImpuestoDescontable, ImpuestosDevolucion,  how='outer', indicator=False).fillna(value=0)

        # leer base de datos de usuarios
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_DB)
        downloader = blob_client.download_blob()
        dbUsers = pd.read_excel(downloader.readall(), sheet_name="Sheet1", header=0)#,engine='openpyxl')
        
        # Construir tabla
        Formato1005 = BuscarId(ImpuestoUnificado,dbUsers)

        dv = Formato1005.apply(lambda x: BuscarDV(x["Numero de identificacion"] if x['Tipo de documento']==31 else 0),axis=1)

        Formato1005.insert(2,"DV", dv)

        Formato1005["Razón social"]= np.where(Formato1005["Primer apellido del informado"]!="","",Formato1005["Razón social"])
        Formato1005["Impuesto Descontable"] = Formato1005["Impuesto Descontable"].apply(np.ceil)
        Formato1005["IVA resultante devoluciones"] = Formato1005["IVA resultante devoluciones"]

        formatoSave = {
            "impuesto":Formato1005["Impuesto Descontable"].sum(),
            "iva":Formato1005["IVA resultante devoluciones"].sum()
            }
        saveToBD(formatoSave,idEjecucion,idProcedencia,Totales)
        
        PutColorsAnsSaveToBlob(Formato1005,container_name)

    except Exception as e:
        dicToReturn = {"error":f'{e}'}
        return json.dumps(dicToReturn)
    dicToReturn = {
        "error":"ninguno",
        "ruta":f'https://{account_name}.blob.core.windows.net/{container_name}/{blob_name_to_save}'
        }
    return json.dumps(dicToReturn)


def ObtenerImpuesto(df,ColumnaValorIngreso):
    impuestoDescontable= df[(df['NumeroCuenta']>=240832)&(df['NumeroCuenta']<=240850)&(df['NIT']!="800197268")&(df['NIT']!="")].iloc[:,ColumnaValorIngreso].sum()
    ivaDescontable= df[(df['NumeroCuenta']>=240820)&(df['NumeroCuenta']<=240820)&(df['NIT']!="800197268")&(df['NIT']!="")].iloc[:,ColumnaValorIngreso].sum()

    print('Total Impuesto Descontable :',impuestoDescontable)
    print('Total Iva Descontable :',ivaDescontable)
    dicToReturn = {
    "Total Impuesto Descontable":impuestoDescontable,
    "Total Iva Descontable":ivaDescontable}
    return (dicToReturn)

def saveToBD(valorHaxa,idEjecucion,idProcedencia,valorContable):
    """
    Almacena las comprobaciones en BD.
    Args:
        valorHaxa (int): Valor total de la suma del formato 1008.
        idEjecucion (int): Número del registro de la ejecución.
        idProcedencia (int): Numero del registro de la procedencia.
        rentaPath (str): Url del archivo pdf de la renta.
    Returns:
        None.
    """

    server = 'rbhaxa.database.windows.net' 
    database = 'haxa' 
    username = 'rbitse' 
    password = 'QbLnBh29XUrDpzX'
    driver= '{ODBC Driver 17 for SQL Server}'# {SQL Server}
    cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server}'+';DATABASE='+database+';ENCRYPT=yes;UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
        
    insert_stmt = (
                    "INSERT INTO Diferencias (id_ejecuccion, id_procedencia, nombre_diferencia,comprobacion, numeroc, observaciones,valor_HAXA) \
                    VALUES (?,?,?,?,?,?,?)"
                    )
    data = (idEjecucion,idProcedencia,"Impuesto Descontable",valorContable['Total Impuesto Descontable'],"2408","",valorHaxa['impuesto'])
    # insertar registro en bd
    cursor.execute(insert_stmt, data)
    data = (idEjecucion,idProcedencia,"Iva Descontable",valorContable['Total Iva Descontable'],"2408","",valorHaxa['iva'])
    # insertar registro en bd
    cursor.execute(insert_stmt, data)
    cnxn.commit()
    return None

def PutColorsAnsSaveToBlob(Datos,container_name):
    '''
    Guarda el archivo excel coloreado del formato en el Blob Storage.
    Args:
        Datos (df): df del formato a guardar.
        container_name (str): Nombre del contenedor en el Blob Storage.
        account_name (str): Nombre de la cuenta del Blob.
        account_key (str): clave de la cuenta del blob.
        blob_name_to_save (str): Nombre del archivo a guardar.
    Returns:
        None.
    '''
    fillOrange = PatternFill(patternType='solid', fgColor='FCBA03')
    fillRed = PatternFill(patternType='solid', fgColor='EE1111')
    
    nombreHojaBalance= "1005" # Nombre de la hoja a guardar
    blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
    buffer = BytesIO()
    excel_buf = Datos.to_excel(sheet_name=nombreHojaBalance,excel_writer=buffer,index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Agrega color a las celdas segun las siguientes normas
    for r in dataframe_to_rows(Datos, index=False, header=True):
        ws.append(r)
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    table = openpyxl.worksheet.table.Table(ref='A1:J'+str(ws.max_row), displayName='Formato1005', tableStyleInfo=mediumStyle)
    ws.add_table(table)

    # Agrega color a las celdas segun las siguientes normas
    for row, rowval in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in rowval:
            
            # rojo si Tipo de doc = NE, no encontrado en bd usuarios
            if f"{cell.column_letter}"=="A" and (f"{cell.value}"=="NE" or f"{cell.value}"=="" or f"{cell.value}"=="None" or f"{cell.value}"=="nan"):
                ws[f"{cell.column_letter}{cell.row}"].fill = fillRed

    # Subir excel a blob
    abs_container_client = blob_service_client.get_container_client(container=container_name)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        blockBlob = abs_container_client.upload_blob(name=blob_name_to_save,data=stream)

    return None

def separarCuentas(df):
    '''
    Separa nits y la razón social del número de cuenta.
    Args:
        df (df): df del balance.

    Returns:
        df: df con 3 columnas extras (Razón social, NIT y NumeroCuenta).
    '''
    cuentas=df["Cuentas"]
    if "Descripción" in cuentas[0]:
        cuentas=cuentas[1:]
        df = df[1:]
    VectorCuentas = []
    VectorNits = []
    VectorNombres = []
    for cuenta in cuentas:  
        try:
            nit,name = cuenta.split(maxsplit=1)
        except Exception:
            try:
                cta = int(str(cuenta).strip().replace('.', ''))
            except Exception:
                cta = 0
            name = ""
            nit = ""
        VectorNombres.append(name)
        VectorCuentas.append(cta)
        VectorNits.append(nit)
    df.insert(1,"Razón social",VectorNombres)
    df.insert(1,"NIT",VectorNits)
    df.insert(1,"NumeroCuenta",VectorCuentas)
    return df


def UnificarClientesPorCuenta(df,limiteInferiorCta,LimiteSuperiorCta,nombreDatoColumna,ColumnaValorImpuesto):
    '''
    Obtiene el valor de la columnaValorIngreso del dfBalance para terceros únicos entre un rango definido.
    Args:
        df (df): df del balance con el número de cuenta separada de la información del tercero.
        limiteInferiorCta (int): rango mínimo del valor de cta a buscar, incluido
        LimiteSuperiorCta (int): rango máximo del valor de la busqueda, no incluido 
        nombreDatoColumna (str): nombre de la columna a insertar en el df de salida
        ColumnaValorIngreso (int): Número de la columna Saldo final a (del balance) 

    Returns:
        datosPorCliente: df con el Numero de identificacion, Razón social y el valor contable del tercero.
    '''
    datosPorCliente = pd.DataFrame()
    listaSoloClientesImpuesto = df[(df['NIT']!="")&(df['NumeroCuenta']>=limiteInferiorCta)&(df['NumeroCuenta']<=LimiteSuperiorCta)&(df['NIT']!="800197268")]
    listaUnica = listaSoloClientesImpuesto["Razón social"].unique().tolist()
    VectorNombres = []
    VectorIdentificacion = []
    VectorTotales = []
    for cliente in listaUnica:
        total = listaSoloClientesImpuesto[(listaSoloClientesImpuesto['Razón social']==cliente)].iloc[:,ColumnaValorImpuesto].sum()
        VectorIdentificacion.append(df[df['Razón social']==cliente]['NIT'].iloc[0])
        VectorNombres.append(cliente)
        VectorTotales.append(total)
    datosPorCliente.insert(0,"Razón social",VectorNombres)
    datosPorCliente.insert(0,"Numero de identificacion",VectorIdentificacion)
    datosPorCliente.insert(2,nombreDatoColumna,VectorTotales)
    datosPorCliente=datosPorCliente[datosPorCliente[nombreDatoColumna]!=0]
    # datosPorCliente[nombreDatoColumna] = datosPorCliente[nombreDatoColumna]*-1
    return datosPorCliente

def BuscarId(dfBalance,bd):
    '''
    Agrega la información del tercero.
    Args:
        dfBalance (df): df de terceros por cada concepto.
        db (df): df de la base de datos de los usuarios.
        dbMupios (df): df de la bd de los municipios de Colombia.
        
    Returns:
        dfBalance: dfBalance con información del tercero correspondiente.
    '''
    TiposDoc={'C':13,'E':31,'N':31,'O':43,'X':31}
    ListadoIds = dfBalance["Numero de identificacion"].tolist()
    vectorCoincidencias = []
    vectorPrimerApellido = []
    vectorSegundoApellido = []
    vectorNombre = []
    vectorOtrosNombres = []
    # vectorDV = []
    bd['Id']=bd['Código'].apply(lambda x: x.strip().split("-")[0])
    for Id in ListadoIds:
        # if len(bd[(bd['Id']==Id)]['Código'].iloc[0].split("-"))==2:
        #     vectorDV.append(bd[(bd['Id']==Id)]['Código'].iloc[0].strip().split("-")[1])
        # else:
        #     vectorDV.append("")
        try:
            tipoId = bd[(bd['Id']==Id)]['Tipo de identificación'].iloc[0]
            NumeroTipoID = TiposDoc[tipoId]
            if NumeroTipoID == 13:
                nombrecompleto = dfBalance[(dfBalance["Numero de identificacion"]==str(Id))]['Razón social'].iloc[0]
                if len(nombrecompleto.split())==1:
                    vectorPrimerApellido.append(nombrecompleto)
                    vectorSegundoApellido.append("")
                    vectorNombre.append("")
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())==2:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append("")
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())==3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[2])
                    vectorOtrosNombres.append("")
                elif len(nombrecompleto.split())>3:
                    vectorPrimerApellido.append(nombrecompleto.split()[0]+" "+nombrecompleto.split()[1] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[0]+" "+nombrecompleto.split()[1]+" "+nombrecompleto.split()[2] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[0])
                    vectorSegundoApellido.append(nombrecompleto.split()[2] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[3] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[1])
                    vectorNombre.append(nombrecompleto.split()[3] if nombrecompleto.split()[0].lower()=="del" 
                                            else nombrecompleto.split()[4] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" 
                                            else nombrecompleto.split()[2])
                    vectorOtrosNombres.append(nombrecompleto.split(maxsplit=4)[4] if nombrecompleto.split()[0].lower()=="del"
                                            else "" if len(nombrecompleto.split())==5 and nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la"
                                            else nombrecompleto.split(maxsplit=5)[5] if nombrecompleto.split()[0].lower()=="de" and nombrecompleto.split()[1].lower()=="la" and len(nombrecompleto.split())>5
                                            else nombrecompleto.split(maxsplit=3)[3])
            else:
                vectorPrimerApellido.append("")
                vectorSegundoApellido.append("")
                vectorNombre.append("")
                vectorOtrosNombres.append("")            
            vectorCoincidencias.append(TiposDoc[tipoId])
        except Exception:
            vectorCoincidencias.append("NE")  
            vectorPrimerApellido.append("")
            vectorSegundoApellido.append("")
            vectorNombre.append("")
            vectorOtrosNombres.append("")

    dfBalance.insert(0,'Tipo de documento', vectorCoincidencias)
    dfBalance.insert(2,'Otros nombres del informado', vectorOtrosNombres)
    dfBalance.insert(2,'Primer nombre del informado', vectorNombre)
    dfBalance.insert(2,'Segundo apellido del informado', vectorSegundoApellido)
    dfBalance.insert(2,'Primer apellido del informado', vectorPrimerApellido)
    # dfBalance.insert(2,'DV', vectorDV)

    return dfBalance

def BuscarDV(nit):
    """
    Calcula el dígito de verificación de un número de identificación tributaria (NIT) colombiano.
    Args:
        nit (str, int): Número de identificación tributaria.

    Returns:
        dv: Dígito de verificación.
    """
    nit = str(nit).strip()
    if nit.isnumeric() and len(nit) > 2:
        res = sum(int(digit) * weight for digit, weight in zip(reversed(nit), [3, 7, 13, 17, 19, 23, 29, 37, 41, 43, 47, 53, 59, 67, 71]))
        dv = 0 if res%11 == 0 else 1 if res%11 == 1 else int(11-res%11)
    else: dv = None
    return dv
