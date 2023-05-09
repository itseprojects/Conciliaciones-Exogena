# -*- coding: utf-8 -*-
"""
@author: HaroldFerneyGomez
"""
from datetime import date
import json
import logging
import unicodedata
from azure.storage.blob import BlobSasPermissions, generate_blob_sas, BlobServiceClient
import pandas as pd
import azure.functions as func
import openpyxl
from openpyxl.styles import PatternFill
from tempfile import NamedTemporaryFile
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from io import BytesIO
import pyodbc
import re

"""
Formatos para extractos e inversiones
Bancos de cuentas que inician por 1110 y 1120
inversiones de cuenta 12 que sean diferentes a 0

Falta inversiones, solicitar un balance que contenga estos datos para crear la funcionalidad
"""
# Datos del blob storage
account_name = 'itseblobdev' #'probepython'
account_key = 'd9sOh0WeqvVF66NQnyWKZWFL/KDje0LizX8UyFWpWX39lLX2C8fxnqRtYD2lOFvNp6aaayQsAq7T+AStvsHyew==' #Zas0npJX9ryEm4hmW/gatWr8aI91oOCvt+qbQKqWrZJCmhv5qh6S/w6ittYYaDBRjRnoxa0h+A8H+ASttcvrrQ=='

blob_name_to_save = 'Formato1012-'+ str(date.today())+'.xlsx' # Archivo excel a guardar en el blob correspondiente al formato
MupiosenBlob = "Municipios_de_Colombia.xlsx"

import requests

def BuscarDV(nit):
    # """Calcula el dígito de verificación de un número de identificación tributaria (NIT) colombiano.    Args:        nit (str, int): Número de identificación tributaria.    Returns:        dv: Dígito de verificación.    """    
    nit = str(nit).strip()
    if nit.isnumeric() and len(nit) > 2:
        res = sum(int(digit) * weight for digit, weight in zip(reversed(nit), [3, 7, 13, 17, 19, 23, 29, 37, 41, 43, 47, 53, 59, 67, 71]))
        dv = 0 if res%11 == 0 else 1 if res%11 == 1 else int(11-res%11)
    else: 
        dv = None    
    return dv
'''
def get_extract_value(bank, pdf_path, password=None):
    if bank == "BBVA":
        a, b, c, d = 270, 340, 300, 100
        row_name = 'SALDO  FINAL'   

    elif bank == "DAVIVIENDA":
        a, b, c, d = 190, 90, 300, 100
        row_name = 'Nuevo Saldo'

    elif bank == "BANCOLOMBIA":
        a, b, c, d = 300, 10, 280, 80
        row_name = 'SALDO ACTUAL'

    # elif bank == "CTA CORRIENTE BANCOLOMBIA":
    #     a, b, c, d = 300, 10, 280, 80
    #     row_name = 'SALDO ACTUAL'
    else:
        return False
    try:
        ext = read_pdf(
            pdf_path, 
            pages='all', 
            area=(a, b, a+d, b+c)
        )
    except:
        if password:
            try:
                ext = read_pdf(
                    pdf_path, 
                    pages='all', 
                    password=password,
                    area=(a, b, a+d, b+c)
                )
            except:
                dv = BuscarDV(password)
                ext = read_pdf(
                    pdf_path, 
                    pages='all', 
                    password=password+str(dv),
                    area=(a, b, a+d, b+c)
                )

    index = ext[0][ext[0].columns[0]].loc[lambda x: x==row_name].index[0]
    return str(ext[0][ext[0].columns[-1]].loc[index]).replace("$","").replace(",","")
'''
def main(req: func.HttpRequest) -> func.HttpResponse:
    Cliente = req.get_json().get('Cliente') # cliente1
    TipoBalance = req.get_json().get('TipoBalance') #"SIESA"
    blob_name_DB = req.get_json().get('BaseDeDatos') #'BASE DE DATOS EXELTIS.xls'
    balanceFile = req.get_json().get('Balance')
    idEjecucion = req.get_json().get('IdEjecucion')
    idProcedencia = req.get_json().get('IdProcedencia')
    extract = req.get_json().get('Extractos')
    # rentaPath = req.get_json().get('RentaUrl')
    # planillas = req.get_json().get('Planillas')
    
    # dfPlanillas = analizarPlanillas(planillas,Cliente)
    if TipoBalance=="SIESA":
        HeaderHojaDB = 0
        nombreHojaDB="Sheet1"
        exito = WorkSiesa(Cliente,balanceFile,blob_name_DB,idEjecucion,idProcedencia,extract)#,rentaPath)
    else: 
        dicToReturn = {"error":"Tipo de balance no implementado"}
        exito = json.dumps(dicToReturn) 

    return func.HttpResponse(f"{exito}", status_code=200 )

def getData(urlPDF,cliente,tipoPlanilla,password=""):
    """
    Función que recibe la URL de un archivo PDF, el nombre del cliente y el tipo de planilla y devuelve el nombre del archivo procesado.
    
    Parámetros:
    - urlPDF (str): URL del archivo PDF a procesar.
    - cliente (str): Nombre del cliente que se está procesando.
    - tipoPlanilla (str): Tipo de planilla que se está procesando.
    
    Retorna:
    - (str): Ruta del archivo procesado.
    """
    url = "https://readpdfs.azurewebsites.net/api/readpdf"
    data = {
        "tipoPDF":tipoPlanilla,
        "id":3,
        "urlPDF":urlPDF,
        "passwordPDF":password,
        "cliente": cliente
    }
    r  = requests.post(url,json=data)

    return str(r.json()['valor'])

def WorkSiesa(container_name,balanceFile,blob_name_DB,idEjecucion,idProcedencia,extractos=[]):#,rentaPath):  
    '''
    Elabora y guarda el formato 1009 en el Blob Storage, almacena las comprobaciones en BD.
    Args:
        container_name (str): Nombre del contenedor - cliente.
        balanceFile (str): Nombre del blob para el balance.
        blob_name_DB (str): Nombre del blob para la BD de Terceros.
        idEjecucion (int): Número del registro de la ejecución.
        idProcedencia (int): Numero del registro de la procedencia.
        rentaPath (str): Url del archivo pdf de la renta.
    Returns:
        dicToReturn: json con información de la ejecución de la función.
    '''
    HeaderHojaBalance=11
    nombreHojaBalance= "Hoja 1" 
    try:
        # Ingreso al blob storage
        blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
        
        try: # Eliminación de archivo de salida si ya existe en el blob
            blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_to_save)
            blob_client.delete_blob()
        except:
            pass
        
        # leer balance y extraer columna del saldo total
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, nrows=0, header=10,engine='openpyxl')
        ColumnaValorIngreso = Datos.columns.get_loc("Saldo final a")
        
        # leer resto del balance
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = balanceFile)
        downloader = blob_client.download_blob()
        Datos = pd.read_excel(downloader.readall(), sheet_name=nombreHojaBalance, header=HeaderHojaBalance,engine='openpyxl')
        Datos = Datos[~Datos['Cuentas'].isnull()]
        Datos = Datos.drop(Datos.columns[-2:], axis=1)

        Datos1 = Datos[(Datos[Datos.columns[0]].fillna("0").str.contains(r"1110")) & ( (Datos[Datos.columns[1]].fillna("0").str.contains(r"CTA")) | (Datos[Datos.columns[1]].fillna("0").str.contains(r"CUENTA")))].dropna().reset_index()
        Datos2 = Datos[(Datos[Datos.columns[0]].fillna("0").str.contains(r"1120")) & ( (Datos[Datos.columns[1]].fillna("0").str.contains(r"CTA")) | (Datos[Datos.columns[1]].fillna("0").str.contains(r"CUENTA")))].dropna().reset_index()
        Datos = pd.concat([Datos1,Datos2],ignore_index=True).reset_index()
        # Datos = Datos.drop([0,len(Datos.index)-1])
        Datos = Datos.drop(columns=['index','level_0'])
        """
        TercerosPorConcepto = TercerosPorConcepto.drop(TercerosPorConcepto[(TercerosPorConcepto['Ingresos Brutos Recibidos']==0)].index)

        TercerosPorConcepto.rename(columns = {'Ingresos Brutos Recibidos':'Saldo a CXC a 31 de diciembre'}, inplace = True)
        
        saveToBD(TercerosPorConcepto["Saldo a CXC a 31 de diciembre"].sum(),idEjecucion,idProcedencia,DatosSeparados[(DatosSeparados['NumeroCuenta']==2)].sum()[ColumnaValorIngreso]*-1)#,rentaPath)    
        
        """
        Datos = Datos.rename(columns= {Datos.columns[-1]:"Saldo"})
        Datos["Saldo"] = Datos["Saldo"].astype(float)
        pattern_order=re.compile(r"\b(CTA|CORRIENTE|AHORROS|[0-9\-]+)")
        Datos["NombreBanco"] = Datos.iloc[:, 1].apply(lambda x: re.sub(pattern_order, "", x).lower().replace(" ", ""))
        Datos = Datos.groupby("NombreBanco")["Saldo"].sum().reset_index()
        extt = []
        for index, row in Datos.iterrows():
            extb = {}
            correct_value = 0
            bank = ""
            for extract in extractos:
                if extract["bank"].lower().__contains__('itau'):
                    pass
                elif extract["bank"].lower() in str(row["NombreBanco"]):
                    extract_value = 0.0
                    for dato in extract["datos"]:
                        extract_value += float(getData(dato["path"], container_name, "extracto"+str(extract["bank"]).lower().strip(),dato["password"]))
                    if float(extract_value) != float(row['Saldo']):
                        correct_value = float(extract_value)
                    bank = extract["bankname"]

            if bank:
                extb["Razón social informado"] = bank
            else:
                #pattern_order = r'[0-9]'
                extb["Razón social informado"] = row['NombreBanco']

            if correct_value == 0:
                extb["Vr a 31 de diciembre"] = float(row['Saldo'])
            else:
                extb["Vr a 31 de diciembre"] = correct_value
                correct_value, bank = 0, ""
            
            extb["NombreBanco"] = row["NombreBanco"]
            extb["Concepto"] = "1110"
            
            extt.append(extb)

            df = pd.DataFrame(extt)
            agg_ = {
                'NombreBanco':'first',
                'Concepto':'first',
                'Vr a 31 de diciembre':'sum'
            }
            df = df.groupby('Razón social informado', as_index=False, sort=False).agg(agg_)
            # df.to_excel("sample_.xlsx", index=None)
        
        df = pd.DataFrame(extt)
        agg_ = { 
            'NombreBanco':'first',
            'Concepto':'first',
            'Vr a 31 de diciembre':'sum'
        }
        df = df.groupby('Razón social informado', as_index=False, sort=False).agg(agg_)
        # leer base de datos de usuarios
        blob_client = blob_service_client.get_blob_client(container = container_name, blob = blob_name_DB)
        downloader = blob_client.download_blob()
        dbUsers = pd.read_excel(downloader.readall(), sheet_name="Sheet1", header=0)#,engine='openpyxl')
        
        df = BuscarId(df,dbUsers)

        dv = df.apply(lambda x: BuscarDV(x["Número identificación del informado"] if x['Tipo documento']==31 else 0),axis=1)
        df.insert(3,"DV", dv)
        df = df.sort_values(by=['Concepto','Número identificación del informado'],ascending=True)
        df = df.reindex(columns=['Concepto',
            'Tipo documento',
            'Número identificación del informado',
            'DV',
            'Primer apellido del informado',
            'Segundo apellido del informado',
            'Primer nombre del informado',
            'Otros nombres del informado',
            'Razón social informado',
            'País residencia',
            'Vr a 31 de diciembre'])
        #df.to_excel(f"{req_body.get('path_final')}sample_.xlsx", index=None)
        # Ajustar parte estética y almacenar en el BLob Storage
        df = df.drop(df[(df['Vr a 31 de diciembre']==0) ].index)
        df['Vr a 31 de diciembre'] = df['Vr a 31 de diciembre'].apply(np.ceil)
        df['Razón social informado'] = df['Razón social informado'].apply(lambda x: next((bancos['bankname'] for bancos in extractos if bancos['bank'] in x), x))
        PutColorsAnsSaveToBlob(df,container_name)
        dicToReturn = {
            "error":"ninguno",
            "ruta":f'https://{account_name}.blob.core.windows.net/{container_name}/{blob_name_to_save}'
            }

    except Exception as e:
        dicToReturn = {"error":f"{e}"}
    return json.dumps(dicToReturn)

    

"""
def saveToBD(valorHaxa,idEjecucion,idProcedencia,valorContable):#,rentaPath):
    '''
    Almacena las comprobaciones en BD.
    Args:
        valorHaxa (int): Valor total de la suma del formato 1008.
        idEjecucion (int): Número del registro de la ejecución.
        idProcedencia (int): Numero del registro de la procedencia.
        rentaPath (str): Url del archivo pdf de la renta.
    Returns:
        None.
    '''
    # leer pdf renta renglon 38
    # area = (24, 22, 622, 779)
    # df = tabula.read_pdf(rentaPath, area=[262,195,272,311], stream=False, pages=1, pandas_options={'header': None})
    # dif = int(str(df[0].iloc[0,0]).replace(",",""))
    # comprobations = pd.DataFrame(columns=['valorPDF'])
    # comprobations = comprobations.append({"valorPDF":dif},ignore_index=True)
    # diferencia = "" if dif==valorHaxa else "Ajuste al peso" if abs(dif-int(valorHaxa))<1000 else "Revisar valores negativos del formato 1008"
    # comprobations.insert(0,"diferencia",diferencia)
    
    server = 'rbhaxa.database.windows.net' 
    database = 'haxa' 
    username = 'rbitse' 
    password = 'QbLnBh29XUrDpzX'
    driver= '{ODBC Driver 17 for SQL Server}'# {SQL Server}
    cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server}'+';DATABASE='+database+';ENCRYPT=yes;UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    
    # condicion = f'id_ejecuccion = {idEjecucion} AND id_procedencia = {idProcedencia}'
    # sql = "DELETE FROM Diferencias" + " WHERE " + condicion
    # cursor.execute(sql)
    # cnxn.commit()
      
    insert_stmt = (
                    "INSERT INTO Diferencias (id_ejecuccion, id_procedencia, nombre_diferencia,comprobacion, numeroc, observaciones,valor_HAXA) \
                    VALUES (?,?,?,?,?,?,?)"
                    )
    data = (idEjecucion,idProcedencia,"Total Contable",valorContable,"2","",valorHaxa)
    # insertar registro en bd
    cursor.execute(insert_stmt, data)
    cnxn.commit()
    return None
"""

def BuscarId(dfBalance,bd):
    '''
    Agrega la información del tercero.
    Args:
        dfBalance (df): df de terceros por cada concepto.
        db (df): df de la base de datos de los usuarios.
        dbMupios (df): df de la bd de los municipios de Colombia.
        
    Returns:
        dfBalance: dfBalance con información del tercero correspondiente.
    '''
    TiposDoc={'C':13,'E':31,'N':31,'O':43,'X':31}   # bd asigna estas letras según el tipo de doc, se asignan # según normativa del formato
    
    dfBalance['Razón social informado uni'] = [unicodedata.normalize('NFKD', x).encode('ASCII', 'ignore').decode("UTF-8") for x in dfBalance['NombreBanco']] # Razón social informado']]
    ListadoIds = dfBalance["Razón social informado uni"].tolist()
    vectorTipoId = []
    vectorNumeroId = []
    vectorRazonSocial = []
    vectorPais = []

    bd['Razón social unicode'] = [unicodedata.normalize('NFKD', x).encode('ASCII', 'ignore').decode("UTF-8") for x in bd['Razón social']]
    try:
        bd['Pais'] = [unicodedata.normalize('NFKD', x).encode('ASCII', 'ignore').decode("UTF-8") for x in bd['Pais']]
    except:
        pass
    for Nombre in ListadoIds:
        # Buscar si existe el tercero en la db y agrega su información, si no existe agrega NE a columna tipo de documento
        try:
            firstMpio = bd.loc[bd['Razón social unicode'].str.contains(Nombre,case=False) & bd['Razón social'].str.contains('banco', case=False)].sort_values('Razón social')
            if firstMpio.empty:
                firstMpio = bd.loc[bd['Razón social unicode'].str.contains(Nombre,case=False)].sort_values('Razón social')
            # Si no encuentra la ciudad o el tipo de identificación es X para el tercero, probablemente no sea de Colombia
            if firstMpio.empty:
                firstMpio=pd.Series({'Tipo de tercero':'',	'Tipo de identificación':'NE',	'Numero identificación':'',
                                    'Razón social':f'{Nombre}',	'Pais':''})
        #     # Selecciona el primer municipio de firstMpio
            else: firstMpio = firstMpio.iloc[0,:]
            
            tipoId = firstMpio['Tipo de identificación']
            vectorTipoId.append(TiposDoc[tipoId])
            vectorNumeroId.append(firstMpio['Numero identificación'].strip().split("-")[0])
            vectorRazonSocial.append(firstMpio['Razón social'])
            try:
                if str(firstMpio['Pais']).lower()=="colombia":
                    vectorPais.append(169)
                else: vectorPais.append('')
            except:
                vectorPais.append("")
            

        except Exception:
            vectorTipoId.append("NE")
            vectorNumeroId.append('')
            vectorRazonSocial.append(Nombre)
            vectorPais.append("")
        
    dfBalance.insert(1,'Tipo documento', vectorTipoId)
    dfBalance.insert(1,'Número identificación del informado', vectorNumeroId)
    dfBalance.insert(1,'Otros nombres del informado', "")
    dfBalance.insert(1,'Primer nombre del informado', "")
    dfBalance.insert(1,'Segundo apellido del informado', "")
    dfBalance.insert(1,'Primer apellido del informado', "")
    dfBalance['Razón social informado']=vectorRazonSocial
    dfBalance.insert(1,'País residencia',vectorPais)

    return dfBalance

def PutColorsAnsSaveToBlob(Datos,container_name):
    '''
    Guarda el archivo excel coloreado del formato en el Blob Storage.
    Args:
        Datos (df): df del formato a guardar.
        container_name (str): Nombre del contenedor en el Blob Storage.
        account_name (str): Nombre de la cuenta del Blob.
        account_key (str): clave de la cuenta del blob.
        blob_name_to_save (str): Nombre del archivo a guardar.
    Returns:
        None.
    '''
    # Asigna el color rojo para advertir sobre datos relacionados de la base de datos de usuarios y amarillo para datos relacionados con el balance
    fillOrange = PatternFill(patternType='solid', fgColor='FCBA03')
    fillRed = PatternFill(patternType='solid', fgColor='EE1111')
    nombreHojaBalance= "iva+Names" # Nombre de la hoja a guardar
    blob_service_client = BlobServiceClient(account_url = f'https://{account_name }.blob.core.windows.net/', credential = account_key)
    buffer = BytesIO()
    excel_buf = Datos.to_excel(sheet_name=nombreHojaBalance,excel_writer=buffer,index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(Datos, index=False, header=True):
        ws.append(r)
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    table = openpyxl.worksheet.table.Table(ref='A1:K'+str(ws.max_row), displayName='Formato1001ByHG', tableStyleInfo=mediumStyle)
    ws.add_table(table)
    
    # Agrega color a las celdas segun las siguientes normas
    for row, rowval in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in rowval:
            
            # rojo si Tipo de doc = NE, no encontrado en bd usuarios
            if f"{cell.column_letter}"=="B" and (f"{cell.value}"=="NE" or f"{cell.value}"==""):
                ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
                
            # rojo si Direccion, departamente, municipio o país está vacio
            elif f"{cell.column_letter}"=="C" or f"{cell.column_letter}"=="J" :
                if f"{cell.value}"=="" or f"{cell.value}"=="None" or f"{cell.value}"=="nan":
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillRed
                    
            # Naranja si pagos o abonos en cuenta menores a 500 mil pesos y diferentes de cero
            elif f"{cell.column_letter}"=="K": 
                if cell.value<0:
                    ws[f"{cell.column_letter}{cell.row}"].fill = fillOrange
                    
    # Subir excel al blob Storage
    abs_container_client = blob_service_client.get_container_client(container=container_name)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        blockBlob = abs_container_client.upload_blob(name=blob_name_to_save,data=stream)
    return None
"""
def separarCuentas(df):
    '''
    Separa nits y la razón social del número de cuenta.
    Args:
        df (df): df del balance.

    Returns:
        df: df con 3 columnas extras (Razón social, NIT y NumeroCuenta).
    '''
    cuentas=df["Cuentas"]
    if "Descripción" in cuentas[0]:
        cuentas=cuentas[1:]
        df = df[1:]
    VectorCuentas = []
    VectorNits = []
    VectorNombres = []
    for cuenta in cuentas:  
        try:
            nit,name = cuenta.split(maxsplit=1)
        except Exception:
            try:
                cta = int(str(cuenta).strip().replace('.', ''))
            except Exception:
                cta = 0
            name = ""
            nit = ""
        VectorNombres.append(name)
        VectorCuentas.append(cta)
        VectorNits.append(nit)
    df.insert(1,"Razón social",VectorNombres)
    df.insert(1,"NIT",VectorNits)
    df.insert(1,"NumeroCuenta",VectorCuentas)
    return df

def GetClientsByConcept(DatosSeparados, ColumnaValorIngreso,dbUsers):
    '''
    Obtiene los terceros a partir del diccionario conceptos.
    Args:
        DatosSeparados (df): df del balance con el número de cuenta separada de la información del tercero.
        ColumnaValorIngreso (int): Número de la columna donde se encuentran el valor a procesar.

    Returns:
        TercerosPorConcepto: df con información del concepto y valor contable para el tercero.
    '''
    # Definicion de los conceptos para el formato tipo SIESA
    conceptos = {"2201":[220500,230500,231000], 
                "2202":[231500],
                "2203":[210000], # todo del 21
                "2204":[240000,248800], # todo del 24 a nombre de direccion de im Dian salvo el 2488: secretaria de Hacienda
                "2214":[251000], # 2510, buscarla por nombre cesant y empiece por 2
                "2215":[250500,251500,252500,253000], 
                "2207":[],
                "2209":[],
                "2208":[], 
                "2211":[],
                "2212":[],
                "2213":[],
                "2206":[230000,260000,270000], # todo 23 menos 2305-2315    +26 y 27
                "22xx":[254000] # 254000,: usado, si es natural en conc 2215, de lo contrario va para el conc 2214
    } 
    TercerosPorConcepto = pd.DataFrame()
    TercerosPorConcepto['Concepto'] = None
    TercerosPorConcepto['Numero de identificacion']= None
    TercerosPorConcepto['Razón social'] = None
    TercerosPorConcepto['Ingresos Brutos Recibidos'] = None
    # print(TercerosPorConcepto.columns)
    for clave in conceptos:
        soloConcepto = pd.DataFrame()
        soloConcepto['Numero de identificacion']= None
        soloConcepto['Razón social'] = None
        soloConcepto['Ingresos Brutos Recibidos'] = None
        for conc in conceptos[clave]:
            if conc==230000 and clave=="2206":
                # exceptuar las ctas de otros conceptos: 2305,2310,2315
                #data1 = UnificarClientesPorCuenta(DatosSeparados,conc,230500 ,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                data = UnificarClientesPorCuenta(DatosSeparados,231600,240000 ,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                #data = pd.concat([data1, data2], ignore_index=True ).groupby(['Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()
            elif clave=='22xx':
                data = UnificarClientesPorCuenta(DatosSeparados,conc,260000 ,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
            elif conc==240000 and clave=="2204":
                data = UnificarClientesPorCuenta(DatosSeparados,conc,248800 ,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                data['Numero de identificacion'] = 800197268
                data['Razón social'] = "DIRECCION DE IMPUESTOS Y ADUANAS NACIONALES"
            elif conc==248800 and clave=="2204":
                data = UnificarClientesPorCuenta(DatosSeparados,conc,248900 ,"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                data['Numero de identificacion'] = 899999061
                data['Razón social'] = "SECRETARIA DISTRITAL DE HACIENDA"
            else:
                data = UnificarClientesPorCuenta(DatosSeparados,conc,(conc+ (10000 if str(conc).endswith("0000") else 100 if str(conc).endswith("00") else 1)),"Ingresos Brutos Recibidos",ColumnaValorIngreso)
                if clave=="2204":
                    data['Numero de identificacion'] = 800197268
                    data['Razón social'] = "DIRECCION DE IMPUESTOS Y ADUANAS NACIONALES"
            # Agrupa los terceros por su #id y razon social y suma sus totales
            soloConcepto = pd.concat([soloConcepto, data], ignore_index=True ).groupby(['Numero de identificacion','Razón social']).sum(numeric_only=True).reset_index()
        # Agrega todos los terceros únicos en el concepto al df de salida
        if ~soloConcepto.empty: 
            soloConcepto.insert(0,'Concepto', clave)
            # Condiciones especiales
            if clave=='22xx':
                soloConcepto['Concepto'] = soloConcepto['Numero de identificacion'].apply(get_tipo_persona,args=(dbUsers,))
                clave = '2015'
            TercerosPorConcepto = pd.concat([TercerosPorConcepto, soloConcepto],ignore_index=True)
    return TercerosPorConcepto  

def get_tipo_persona(numero_identificacion,dbUsers):
    '''
    Busca el número de identificación en el DataFrame dbUsers y retorna la posición 2 del registro correspondiente,
    si es que existe.

    Args:
        numero_identificacion (str): Número de identificación a buscar.

    Returns:
        str: La posición 2 del registro correspondiente en el DataFrame dbUsers, o "No encontrado" si no existe.
    '''
    dbUsers['Código'] = dbUsers['Código'].apply(lambda x: x.strip())
    registro = dbUsers.loc[dbUsers['Código'] == numero_identificacion.strip()]
    persona = registro.iloc[0, 1] if not registro.empty else "No encontrado"
    if persona.upper() == 'PERSONA NATURAL':
        return "2215"
    else: return "2214"

def UnificarClientesPorCuenta(df,limiteInferiorCta,LimiteSuperiorCta,nombreDatoColumna,ColumnaValorIngreso):
    '''
    Obtiene el valor de la columnaValorIngreso del dfBalance para terceros únicos entre un rango definido.
    Args:
        df (df): df del balance con el número de cuenta separada de la información del tercero.
        limiteInferiorCta (int): rango mínimo del valor de cta a buscar, incluido
        LimiteSuperiorCta (int): rango máximo del valor de la busqueda, no incluido 
        nombreDatoColumna (str): nombre de la columna a insertar en el df de salida
        ColumnaValorIngreso (int): Número de la columna Saldo final a (del balance) 

    Returns:
        datosPorCliente: df con el Numero de identificacion, Razón social y el valor contable del tercero.
    '''
    datosPorCliente = pd.DataFrame()
    listaSoloClientesIngreso = df[(df['NIT']!="")&(df['NumeroCuenta']>=limiteInferiorCta)&(df['NumeroCuenta']<LimiteSuperiorCta)]
    listaUnica = listaSoloClientesIngreso["Razón social"].unique().tolist()
    VectorNombres = []
    VectorIdentificacion = []
    VectorTotales = []
    for cliente in listaUnica:
        total = listaSoloClientesIngreso[(listaSoloClientesIngreso['Razón social']==cliente)].iloc[:, ColumnaValorIngreso].sum()
        VectorIdentificacion.append(df[df['Razón social']==cliente]['NIT'].iloc[0])
        VectorNombres.append(cliente)
        VectorTotales.append(total)
    
    datosPorCliente.insert(0,"Razón social",VectorNombres)
    datosPorCliente.insert(0,"Numero de identificacion",VectorIdentificacion)
    datosPorCliente.insert(2,nombreDatoColumna,VectorTotales)
    return datosPorCliente
"""
